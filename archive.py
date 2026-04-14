from flask import Blueprint, render_template, request, jsonify, send_file
from db import get_db_connection
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

archive_bp = Blueprint('archive_bp', __name__)

# Helper to check for optional columns (is_archived / deleted_at) with caching
_archive_table_column_cache = {}

def _table_has_column(conn, table_name, column_name):
    key = f"{table_name}.{column_name}"
    if key in _archive_table_column_cache:
        return _archive_table_column_cache[key]
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS"
                " WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s AND COLUMN_NAME = %s",
                (table_name, column_name)
            )
            row = cur.fetchone()
            if isinstance(row, dict):
                exists = int(row.get('cnt', 0)) > 0
            else:
                exists = bool(row and row[0] > 0)
    except Exception:
        exists = False
    _archive_table_column_cache[key] = exists
    return exists

@archive_bp.route('/archive')
def archive_page():
    return render_template('archive.html')

@archive_bp.route('/api/archived-pcs')
def get_archived_pcs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '', type=str)
    
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            # Choose archive predicate: prefer is_archived, fallback to deleted_at, else return empty
            has_is_arch = _table_has_column(conn, 'pcinfofull', 'is_archived')
            has_deleted_at = _table_has_column(conn, 'pcinfofull', 'deleted_at')
            if has_is_arch:
                base_pred = 'pc.is_archived = 1'
            elif has_deleted_at:
                base_pred = 'pc.deleted_at IS NOT NULL'
            else:
                base_pred = '0=1'

            # Get total count
            if search:
                count_query = f"SELECT COUNT(*) AS total FROM pcinfofull pc LEFT JOIN departments dep ON pc.department_id = dep.department_id WHERE {base_pred} AND (pc.pcid LIKE %s OR pc.pcname LIKE %s OR pc.serial_no LIKE %s OR dep.department_name LIKE %s)"
                cur.execute(count_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'))
            else:
                count_query = f"SELECT COUNT(*) AS total FROM pcinfofull pc WHERE {base_pred}"
                cur.execute(count_query)
            total = cur.fetchone()['total']

            # Get paginated data
            offset = (page - 1) * per_page
            # select deleted_at only if available, otherwise return NULL for compatibility
            select_deleted = 'pc.deleted_at' if has_deleted_at else 'NULL AS deleted_at'
            order_clause = 'ORDER BY pc.deleted_at DESC' if has_deleted_at else 'ORDER BY pc.pcid DESC'
            base_select = f"pc.pcid, pc.pcname, dep.department_name, pc.location, pc.acquisition_cost, pc.date_acquired, pc.accountable, pc.serial_no, pc.municipal_serial_no, pc.status, {select_deleted}"

            if search:
                data_query = f"SELECT {base_select} FROM pcinfofull pc LEFT JOIN departments dep ON pc.department_id = dep.department_id WHERE {base_pred} AND (pc.pcid LIKE %s OR pc.pcname LIKE %s OR pc.serial_no LIKE %s OR dep.department_name LIKE %s) {order_clause} LIMIT %s OFFSET %s"
                cur.execute(data_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', per_page, offset))
            else:
                data_query = f"SELECT {base_select} FROM pcinfofull pc LEFT JOIN departments dep ON pc.department_id = dep.department_id WHERE {base_pred} {order_clause} LIMIT %s OFFSET %s"
                cur.execute(data_query, (per_page, offset))
            pcs = cur.fetchall()
            
            # Convert to list of dicts
            pcs_list = []
            for pc in pcs:
                pcs_list.append({
                    'pcid': pc['pcid'],
                    'pcname': pc['pcname'],
                    'department_name': pc['department_name'],
                    'location': pc['location'],
                    'acquisition_cost': pc['acquisition_cost'],
                    'date_acquired': pc['date_acquired'],
                    'accountable': pc['accountable'],
                    'serial_no': pc['serial_no'],
                    'municipal_serial_no': pc['municipal_serial_no'],
                    'status': 'ARCHIVE',
                    'deleted_at': pc['deleted_at'].strftime('%Y-%m-%d %H:%M:%S') if pc['deleted_at'] else ''
                })
            
            total_pages = (total + per_page - 1) // per_page
            
            return jsonify({
                'data': pcs_list,
                'total': total,
                'page': page,
                'per_page': per_page,
                'total_pages': total_pages
            })
    finally:
        conn.close()


@archive_bp.route('/export-archived-pcs')
def export_archived_pcs():
    """Export archived PCs as a formatted Excel workbook."""
    search = request.args.get('search', '', type=str)

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            has_is_arch = _table_has_column(conn, 'pcinfofull', 'is_archived')
            has_deleted_at = _table_has_column(conn, 'pcinfofull', 'deleted_at')
            if has_is_arch:
                base_pred = 'pc.is_archived = 1'
            elif has_deleted_at:
                base_pred = 'pc.deleted_at IS NOT NULL'
            else:
                base_pred = '0=1'

            select_deleted = 'pc.deleted_at' if has_deleted_at else 'NULL AS deleted_at'
            order_clause = 'ORDER BY pc.deleted_at DESC' if has_deleted_at else 'ORDER BY pc.pcid DESC'
            base_select = f"pc.pcid, pc.pcname, dep.department_name, pc.location, pc.acquisition_cost, pc.date_acquired, pc.accountable, pc.serial_no, pc.municipal_serial_no, pc.status, {select_deleted}"

            if search:
                data_query = f"SELECT {base_select} FROM pcinfofull pc LEFT JOIN departments dep ON pc.department_id = dep.department_id WHERE {base_pred} AND (pc.pcid LIKE %s OR pc.pcname LIKE %s OR pc.serial_no LIKE %s OR dep.department_name LIKE %s) {order_clause}"
                cur.execute(data_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'))
            else:
                data_query = f"SELECT {base_select} FROM pcinfofull pc LEFT JOIN departments dep ON pc.department_id = dep.department_id WHERE {base_pred} {order_clause}"
                cur.execute(data_query)

            rows = cur.fetchall()

        rename_map = {
            'pcid': 'PC ID',
            'pcname': 'PC Name',
            'department_name': 'Department',
            'location': 'Location',
            'acquisition_cost': 'Acquisition Cost',
            'date_acquired': 'Date Acquired',
            'accountable': 'Accountable',
            'serial_no': 'Serial No.',
            'municipal_serial_no': 'Municipal Serial No.',
            'status': 'Status',
            'deleted_at': 'Archived Date'
        }

        if rows:
            df = pd.DataFrame(rows)
        else:
            df = pd.DataFrame(columns=list(rename_map.keys()))

        df.rename(columns=rename_map, inplace=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Archived PCs"

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font = Font(name="Calibri", bold=True, color="1F4E79", size=14)
        subtitle_font = Font(name="Calibri", bold=False, color="4472C4", size=11)
        data_font = Font(name="Calibri", size=11, color="333333")
        thin_border = Border(left=Side(style="thin", color="B4C6E7"), right=Side(style="thin", color="B4C6E7"), top=Side(style="thin", color="B4C6E7"), bottom=Side(style="thin", color="B4C6E7"))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")

        # Title and subtitle
        col_count = len(df.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = ws.cell(row=1, column=1, value="Norzagaray College — Archived PCs")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        subtitle_cell = ws.cell(row=2, column=1, value=f"Exported {datetime.now().strftime('%B %d, %Y %I:%M %p')}")
        subtitle_cell.font = subtitle_font
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 6

        header_row = 4
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 26

        # Data rows
        currency_cols = {"Acquisition Cost"}
        date_cols = {"Date Acquired", "Archived Date"}
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=header_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx, col_name in enumerate(df.columns, start=1):
                val = row_data[col_name]
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border

                if col_name in currency_cols and val != "":
                    cell.number_format = '₱#,##0.00'
                    cell.alignment = right_align
                elif col_name in date_cols and val != "":
                    # leave datetime objects as-is for numeric date formatting
                    cell.number_format = 'YYYY-MM-DD'
                    cell.alignment = center_align
                elif col_name in ("PC ID", "Serial No.", "Municipal Serial No."):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # Auto-width
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name)) + 2
            for row_idx in range(header_row + 1, header_row + 1 + len(df)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 50)

        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{header_row + max(1, len(df))}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_name="Archived_PCs.xlsx", as_attachment=True)
    finally:
        conn.close()


@archive_bp.route('/export-archived-items')
def export_archived_items():
    """Export archived items as Excel workbook."""
    search = request.args.get('search', '', type=str)

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            has_is_arch = _table_has_column(conn, 'devices_full', 'is_archived')
            has_deleted_at = _table_has_column(conn, 'devices_full', 'deleted_at')
            if has_is_arch:
                base_pred = 'df.is_archived = 1'
            elif has_deleted_at:
                base_pred = 'df.deleted_at IS NOT NULL'
            else:
                base_pred = '0=1'

            select_deleted = 'df.deleted_at' if has_deleted_at else 'NULL AS deleted_at'
            order_clause = 'ORDER BY df.deleted_at DESC' if has_deleted_at else 'ORDER BY df.accession_id DESC'
            base_select = f"df.accession_id, df.item_name, df.brand_model, df.serial_no, df.municipal_serial_no, df.acquisition_cost, df.date_acquired, df.accountable, dep.department_name, {select_deleted}"

            if search:
                data_query = f"SELECT {base_select} FROM devices_full df LEFT JOIN departments dep ON df.department_id = dep.department_id WHERE {base_pred} AND (df.accession_id LIKE %s OR df.item_name LIKE %s OR df.brand_model LIKE %s OR df.serial_no LIKE %s OR df.municipal_serial_no LIKE %s OR dep.department_name LIKE %s) {order_clause}"
                cur.execute(data_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'))
            else:
                data_query = f"SELECT {base_select} FROM devices_full df LEFT JOIN departments dep ON df.department_id = dep.department_id WHERE {base_pred} {order_clause}"
                cur.execute(data_query)

            rows = cur.fetchall()

        rename_map = {
            'accession_id': 'Accession ID',
            'item_name': 'Item Name',
            'brand_model': 'Brand/Model',
            'serial_no': 'Serial No.',
            'municipal_serial_no': 'Municipal Serial No.',
            'acquisition_cost': 'Acquisition Cost',
            'date_acquired': 'Date Acquired',
            'accountable': 'Accountable',
            'department_name': 'Department',
            'deleted_at': 'Archived Date'
        }

        if rows:
            df = pd.DataFrame(rows)
        else:
            df = pd.DataFrame(columns=list(rename_map.keys()))

        df.rename(columns=rename_map, inplace=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Archived Items"

        # Styles (reuse same palette)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font = Font(name="Calibri", bold=True, color="1F4E79", size=14)
        subtitle_font = Font(name="Calibri", bold=False, color="4472C4", size=11)
        data_font = Font(name="Calibri", size=11, color="333333")
        thin_border = Border(left=Side(style="thin", color="B4C6E7"), right=Side(style="thin", color="B4C6E7"), top=Side(style="thin", color="B4C6E7"), bottom=Side(style="thin", color="B4C6E7"))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")

        col_count = len(df.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = ws.cell(row=1, column=1, value="Norzagaray College — Archived Items")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        subtitle_cell = ws.cell(row=2, column=1, value=f"Exported {datetime.now().strftime('%B %d, %Y %I:%M %p')}")
        subtitle_cell.font = subtitle_font
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 6

        header_row = 4
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 26

        currency_cols = {"Acquisition Cost"}
        date_cols = {"Date Acquired", "Archived Date"}
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=header_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx, col_name in enumerate(df.columns, start=1):
                val = row_data[col_name]
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border

                if col_name in currency_cols and val != "":
                    cell.number_format = '₱#,##0.00'
                    cell.alignment = right_align
                elif col_name in date_cols and val != "":
                    cell.number_format = 'YYYY-MM-DD'
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name)) + 2
            for row_idx in range(header_row + 1, header_row + 1 + len(df)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 50)

        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{header_row + max(1, len(df))}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_name="Archived_Items.xlsx", as_attachment=True)
    finally:
        conn.close()


@archive_bp.route('/export-archived-consumables')
def export_archived_consumables():
    """Export archived consumables as Excel workbook."""
    search = request.args.get('search', '', type=str)

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            has_is_arch = _table_has_column(conn, 'consumables', 'is_archived')
            has_deleted_at = _table_has_column(conn, 'consumables', 'deleted_at')
            if has_is_arch:
                base_pred = 'c.is_archived = 1'
            elif has_deleted_at:
                base_pred = 'c.deleted_at IS NOT NULL'
            else:
                base_pred = '0=1'

            select_deleted = 'c.deleted_at' if has_deleted_at else 'NULL AS deleted_at'
            order_clause = 'ORDER BY c.deleted_at DESC' if has_deleted_at else 'ORDER BY c.accession_id DESC'
            base_select = f"c.accession_id, c.item_name, c.brand, c.quantity, c.unit, dep.department_name, c.location, c.status, {select_deleted}"

            if search:
                data_query = f"SELECT {base_select} FROM consumables c LEFT JOIN departments dep ON c.department_id = dep.department_id WHERE {base_pred} AND (c.accession_id LIKE %s OR c.item_name LIKE %s OR c.brand LIKE %s OR dep.department_name LIKE %s) {order_clause}"
                cur.execute(data_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'))
            else:
                data_query = f"SELECT {base_select} FROM consumables c LEFT JOIN departments dep ON c.department_id = dep.department_id WHERE {base_pred} {order_clause}"
                cur.execute(data_query)

            rows = cur.fetchall()

        rename_map = {
            'accession_id': 'Accession ID',
            'item_name': 'Item Name',
            'brand': 'Brand',
            'quantity': 'Quantity',
            'unit': 'Unit',
            'department_name': 'Department',
            'location': 'Location',
            'status': 'Status',
            'deleted_at': 'Archived Date'
        }

        if rows:
            df = pd.DataFrame(rows)
        else:
            df = pd.DataFrame(columns=list(rename_map.keys()))

        df.rename(columns=rename_map, inplace=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Archived Consumables"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font = Font(name="Calibri", bold=True, color="1F4E79", size=14)
        subtitle_font = Font(name="Calibri", bold=False, color="4472C4", size=11)
        data_font = Font(name="Calibri", size=11, color="333333")
        thin_border = Border(left=Side(style="thin", color="B4C6E7"), right=Side(style="thin", color="B4C6E7"), top=Side(style="thin", color="B4C6E7"), bottom=Side(style="thin", color="B4C6E7"))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")

        col_count = len(df.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = ws.cell(row=1, column=1, value="Norzagaray College — Archived Consumables")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        subtitle_cell = ws.cell(row=2, column=1, value=f"Exported {datetime.now().strftime('%B %d, %Y %I:%M %p')}")
        subtitle_cell.font = subtitle_font
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 6

        header_row = 4
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 26

        numeric_cols = {"Quantity"}
        date_cols = {"Archived Date"}
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=header_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx, col_name in enumerate(df.columns, start=1):
                val = row_data[col_name]
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border

                if col_name in numeric_cols and val != "":
                    cell.number_format = '#,##0'
                    cell.alignment = right_align
                elif col_name in date_cols and val != "":
                    cell.number_format = 'YYYY-MM-DD'
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name)) + 2
            for row_idx in range(header_row + 1, header_row + 1 + len(df)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 50)

        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{header_row + max(1, len(df))}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_name="Archived_Consumables.xlsx", as_attachment=True)
    finally:
        conn.close()

@archive_bp.route('/api/archived-items')
def get_archived_items():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '', type=str)
    
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            # Get total count
            has_is_arch = _table_has_column(conn, 'devices_full', 'is_archived')
            has_deleted_at = _table_has_column(conn, 'devices_full', 'deleted_at')
            if has_is_arch:
                base_pred = 'df.is_archived = 1'
            elif has_deleted_at:
                base_pred = 'df.deleted_at IS NOT NULL'
            else:
                base_pred = '0=1'

            if search:
                count_query = f"SELECT COUNT(*) AS total FROM devices_full df LEFT JOIN departments dep ON df.department_id = dep.department_id WHERE {base_pred} AND (df.accession_id LIKE %s OR df.item_name LIKE %s OR df.brand_model LIKE %s OR df.device_type LIKE %s OR dep.department_name LIKE %s)"
                cur.execute(count_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'))
            else:
                count_query = f"SELECT COUNT(*) AS total FROM devices_full df WHERE {base_pred}"
                cur.execute(count_query)
            total = cur.fetchone()['total']

            # Get paginated data
            offset = (page - 1) * per_page
            select_deleted = 'df.deleted_at' if has_deleted_at else 'NULL AS deleted_at'
            order_clause = 'ORDER BY df.deleted_at DESC' if has_deleted_at else 'ORDER BY df.accession_id DESC'
            base_select = f"df.accession_id, df.item_name, df.brand_model, df.serial_no, df.municipal_serial_no, df.acquisition_cost, df.date_acquired, df.accountable, dep.department_name, {select_deleted}"

            if search:
                data_query = f"SELECT {base_select} FROM devices_full df LEFT JOIN departments dep ON df.department_id = dep.department_id WHERE {base_pred} AND (df.accession_id LIKE %s OR df.item_name LIKE %s OR df.brand_model LIKE %s OR df.serial_no LIKE %s OR df.municipal_serial_no LIKE %s OR dep.department_name LIKE %s) {order_clause} LIMIT %s OFFSET %s"
                cur.execute(data_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', per_page, offset))
            else:
                data_query = f"SELECT {base_select} FROM devices_full df LEFT JOIN departments dep ON df.department_id = dep.department_id WHERE {base_pred} {order_clause} LIMIT %s OFFSET %s"
                cur.execute(data_query, (per_page, offset))
            items = cur.fetchall()
            
            # Convert to list of dicts
            items_list = []
            for item in items:
                items_list.append({
                    'accession_id': item['accession_id'],
                    'item_name': item['item_name'],
                    'brand_model': item['brand_model'],
                    'serial_no': item['serial_no'],
                    'municipal_serial_no': item['municipal_serial_no'],
                    'acquisition_cost': item['acquisition_cost'],
                    'date_acquired': item['date_acquired'],
                    'accountable': item['accountable'],
                    'department_name': item['department_name'],
                    'status': 'ARCHIVED',
                    'deleted_at': item['deleted_at'].strftime('%Y-%m-%d %H:%M:%S') if item['deleted_at'] else ''
                })
            
            total_pages = (total + per_page - 1) // per_page
            
            return jsonify({
                'data': items_list,
                'total': total,
                'page': page,
                'per_page': per_page,
                'total_pages': total_pages
            })
    finally:
        conn.close()


@archive_bp.route('/api/archived-consumables')
def get_archived_consumables():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '', type=str)

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            has_is_arch = _table_has_column(conn, 'consumables', 'is_archived')
            has_deleted_at = _table_has_column(conn, 'consumables', 'deleted_at')
            if has_is_arch:
                base_pred = 'c.is_archived = 1'
            elif has_deleted_at:
                base_pred = 'c.deleted_at IS NOT NULL'
            else:
                base_pred = '0=1'

            if search:
                count_query = f"SELECT COUNT(*) AS total FROM consumables c LEFT JOIN departments dep ON c.department_id = dep.department_id WHERE {base_pred} AND (c.accession_id LIKE %s OR c.item_name LIKE %s OR c.brand LIKE %s OR dep.department_name LIKE %s)"
                cur.execute(count_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'))
            else:
                count_query = f"SELECT COUNT(*) AS total FROM consumables c WHERE {base_pred}"
                cur.execute(count_query)
            total = cur.fetchone()['total']

            offset = (page - 1) * per_page
            select_deleted = 'c.deleted_at' if has_deleted_at else 'NULL AS deleted_at'
            order_clause = 'ORDER BY c.deleted_at DESC' if has_deleted_at else 'ORDER BY c.accession_id DESC'
            base_select = f"c.accession_id, c.item_name, c.brand, c.quantity, c.unit, dep.department_name, c.location, c.status, {select_deleted}"

            if search:
                data_query = f"SELECT {base_select} FROM consumables c LEFT JOIN departments dep ON c.department_id = dep.department_id WHERE {base_pred} AND (c.accession_id LIKE %s OR c.item_name LIKE %s OR c.brand LIKE %s OR dep.department_name LIKE %s) {order_clause} LIMIT %s OFFSET %s"
                cur.execute(data_query, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', per_page, offset))
            else:
                data_query = f"SELECT {base_select} FROM consumables c LEFT JOIN departments dep ON c.department_id = dep.department_id WHERE {base_pred} {order_clause} LIMIT %s OFFSET %s"
                cur.execute(data_query, (per_page, offset))
            rows = cur.fetchall()

            consumables_list = []
            for row in rows:
                consumables_list.append({
                    'accession_id': row['accession_id'],
                    'item_name': row['item_name'],
                    'brand': row['brand'],
                    'quantity': row['quantity'],
                    'unit': row['unit'],
                    'department_name': row.get('department_name'),
                    'location': row['location'],
                    'status': 'ARCHIVED',
                    'deleted_at': row['deleted_at'].strftime('%Y-%m-%d %H:%M:%S') if row['deleted_at'] else ''
                })

            total_pages = (total + per_page - 1) // per_page
            return jsonify({
                'data': consumables_list,
                'total': total,
                'page': page,
                'per_page': per_page,
                'total_pages': total_pages
            })
    finally:
        conn.close()
