from flask import request, render_template, redirect, url_for, flash, Blueprint,jsonify,session,send_file
from db import get_db_connection
from io import BytesIO
import pandas as pd
import pymysql, re
from collections import defaultdict
import uuid
from datetime import date
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from utils.inventory_audit import log_inventory_action
from notification import add_notification

manage_pc_bp = Blueprint('manage_pc_bp', __name__)


def _existing_pc_part_columns(cur):
    """Return list of existing part column names on `pcinfofull` preserving canonical order.

    Uses the provided cursor to query information_schema; returns an ordered
    subset of the known part columns that actually exist in the table.
    """
    parts = ['monitor', 'motherboard', 'ram', 'storage', 'gpu', 'psu', 'casing', 'mouse', 'keyboard', 'other_parts']
    try:
        placeholders = ','.join(['%s'] * len(parts))
        cur.execute(f"SELECT COLUMN_NAME FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'pcinfofull' AND COLUMN_NAME IN ({placeholders})", parts)
        rows = cur.fetchall()
        found = set()
        for r in rows:
            if isinstance(r, dict):
                found.add(r.get('COLUMN_NAME'))
            else:
                found.add(r[0])
        return [p for p in parts if p in found]
    except Exception:
        # If anything goes wrong, conservatively assume common part columns
        # except `mouse` which may be missing on older schemas.
        return [p for p in parts if p != 'mouse']

def _parse_tokens(value):
    if not value:
        return set()
    return set(t.strip() for t in re.split(r"[,;\n]+", str(value)) if t and t.strip())


def _mark_device_values_in_use(cur, value, performed_by=None):
    """Mark devices referenced by `value` as IN USE.
    Supports accession_id, exact serials, and partial case-insensitive name/brand matches.
    """
    if not value:
        return

    tokens = _parse_tokens(value)
    for token in tokens:
        try:
            print(f"[mark_in_use] token={token!r}")
        except Exception:
            pass
        # try accession id (exact)
        try:
            acc_id = int(token)
        except Exception:
            acc_id = None

        if acc_id is not None:
            cur.execute("SELECT accession_id, status, quantity, device_type FROM devices_full WHERE accession_id = %s", (acc_id,))
            row = cur.fetchone()
            if row:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
                old_status = row.get('status') if isinstance(row, dict) else row[1]
                old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
                try:
                    cur_qty = int(old_qty) if old_qty is not None else None
                except Exception:
                    cur_qty = None

                # If quantity is known, decrement by 1; otherwise mark as IN USE
                if cur_qty is None:
                    if old_status != 'IN USE':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                        except Exception:
                            pass
                        try:
                            print(f"[mark_in_use] updated accession_id={aid} {old_status!r} -> 'IN USE'")
                        except Exception:
                            pass
                else:
                    if cur_qty > 0:
                        new_qty = max(cur_qty - 1, 0)
                        cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                        except Exception:
                            pass
                        if new_qty == 0 and old_status != 'IN USE':
                            cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                            try:
                                log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                            except Exception:
                                pass
                        try:
                            print(f"[mark_in_use] accession_id={aid} quantity {cur_qty} -> {new_qty}")
                        except Exception:
                            pass
                    else:
                        if old_status != 'IN USE':
                            cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                            try:
                                log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                            except Exception:
                                pass
                            try:
                                print(f"[mark_in_use] accession_id={aid} set to 'IN USE' (quantity already 0)")
                            except Exception:
                                pass
            else:
                try:
                    print(f"[mark_in_use] accession_id {acc_id} not found")
                except Exception:
                    pass
            continue

        # exact serial / municipal serial (may match multiple rows)
        cur.execute("SELECT accession_id, status, quantity, device_type FROM devices_full WHERE serial_no = %s OR municipal_serial_no = %s", (token, token))
        rows = cur.fetchall()
        if rows:
            try:
                matched = [(r.get('accession_id') if isinstance(r, dict) else r[0]) for r in rows]
                print(f"[mark_in_use] serial match token={token!r} matched_accession_ids={matched}")
            except Exception:
                pass
            for row in rows:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
                old_status = row.get('status') if isinstance(row, dict) else row[1]
                old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
                try:
                    cur_qty = int(old_qty) if old_qty is not None else None
                except Exception:
                    cur_qty = None

                if cur_qty is None:
                    if old_status != 'IN USE':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                        except Exception:
                            pass
                        try:
                            print(f"[mark_in_use] updated accession_id={aid} {old_status!r} -> 'IN USE'")
                        except Exception:
                            pass
                else:
                    if cur_qty > 0:
                        new_qty = max(cur_qty - 1, 0)
                        cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                        except Exception:
                            pass
                        if new_qty == 0 and old_status != 'IN USE':
                            cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                            try:
                                log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                            except Exception:
                                pass
                        try:
                            print(f"[mark_in_use] accession_id={aid} quantity {cur_qty} -> {new_qty}")
                        except Exception:
                            pass
                    else:
                        if old_status != 'IN USE':
                            cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                            try:
                                log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                            except Exception:
                                pass
                            try:
                                print(f"[mark_in_use] accession_id={aid} set to 'IN USE' (quantity already 0)")
                            except Exception:
                                pass
            continue

        # partial, case-insensitive match on item_name or brand_model — pick one representative row
        like = f"%{token.lower()}%"
        cur.execute("SELECT accession_id, status, quantity, device_type FROM devices_full WHERE LOWER(item_name) LIKE %s OR LOWER(brand_model) LIKE %s ORDER BY accession_id LIMIT 1", (like, like))
        row = cur.fetchone()
        if row:
            try:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
                print(f"[mark_in_use] partial-name match token={token!r} matched_accession_id={aid}")
            except Exception:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
            old_status = row.get('status') if isinstance(row, dict) else row[1]
            old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
            try:
                cur_qty = int(old_qty) if old_qty is not None else None
            except Exception:
                cur_qty = None

            if cur_qty is None:
                if old_status != 'IN USE':
                    cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                    except Exception:
                        pass
                    try:
                        print(f"[mark_in_use] updated accession_id={aid} {old_status!r} -> 'IN USE'")
                    except Exception:
                        pass
            else:
                if cur_qty > 0:
                    new_qty = max(cur_qty - 1, 0)
                    cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                    except Exception:
                        pass
                    if new_qty == 0 and old_status != 'IN USE':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                        except Exception:
                            pass
                    try:
                        print(f"[mark_in_use] accession_id={aid} quantity {cur_qty} -> {new_qty}")
                    except Exception:
                        pass
                else:
                    if old_status != 'IN USE':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('IN USE', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'IN USE', performed_by)
                        except Exception:
                            pass
                        try:
                            print(f"[mark_in_use] accession_id={aid} set to 'IN USE' (quantity already 0)")
                        except Exception:
                            pass
        continue


def _is_token_referenced_elsewhere(cur, token, exclude_pcid=None):
    parts = _existing_pc_part_columns(cur)
    if not parts:
        return False
    ex = exclude_pcid or 0
    placeholders = ' OR '.join([f"{p} LIKE %s" for p in parts])
    like = f"%{token}%"
    params = [ex] + [like] * len(parts)
    # Try including is_archived guard if available, fall back if DB lacks the column
    try:
        cur.execute(f"SELECT COUNT(*) AS cnt FROM pcinfofull WHERE pcid != %s AND COALESCE(is_archived, 0) = 0 AND ({placeholders})", params)
    except Exception:
        cur.execute(f"SELECT COUNT(*) AS cnt FROM pcinfofull WHERE pcid != %s AND ({placeholders})", params)
    row = cur.fetchone()
    cnt = row.get('cnt') if isinstance(row, dict) else (row[0] if row else 0)
    return int(cnt or 0) > 0


def _release_device_values_if_unreferenced(cur, value, exclude_pcid=None, performed_by=None):
    """Release devices referenced by `value` back to Available if no other active PC references them."""
    if not value:
        return

    tokens = _parse_tokens(value)
    for token in tokens:
        try:
            print(f"[release] token={token!r}")
        except Exception:
            pass

        # try accession id
        try:
            aid = int(token)
        except Exception:
            aid = None

        if aid is not None:
            cur.execute("SELECT accession_id, status, quantity, device_type FROM devices_full WHERE accession_id = %s", (aid,))
            row = cur.fetchone()
            if not row:
                try:
                    print(f"[release] accession_id {aid} not found")
                except Exception:
                    pass
                continue
            old_status = row.get('status') if isinstance(row, dict) else row[1]
            old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
            try:
                cur_qty = int(old_qty) if old_qty is not None else None
            except Exception:
                cur_qty = None

            if _is_token_referenced_elsewhere(cur, token, exclude_pcid):
                try:
                    print(f"[release] accession_id={aid} still referenced elsewhere; skipping")
                except Exception:
                    pass
                continue

            if cur_qty is None:
                if old_status != 'Available':
                    cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Available', aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Available', performed_by)
                    except Exception:
                        pass
                    try:
                        print(f"[release] updated accession_id={aid} {old_status!r} -> 'Available'")
                    except Exception:
                        pass
            else:
                new_qty = cur_qty + 1
                cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                try:
                    log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                except Exception:
                    pass
                if old_status != 'Available':
                    cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Available', aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Available', performed_by)
                    except Exception:
                        pass
                try:
                    print(f"[release] accession_id={aid} quantity {cur_qty} -> {new_qty} and status -> 'Available'")
                except Exception:
                    pass
            continue


def _mark_device_values_damaged(cur, value, exclude_pcid=None, performed_by=None):
    """Mark devices referenced by `value` as Damaged and return them to Manage Item list.

    Behavior:
      - For single accession rows (quantity is NULL) set `status = 'Damaged'`.
      - For stock/quantity rows (quantity is NOT NULL) increment `quantity` by 1
        (do not change overall row status to avoid marking entire stock as damaged).
      - Skip tokens that are still referenced by other active PCs.
    """
    if not value:
        return

    tokens = _parse_tokens(value)
    for token in tokens:
        try:
            print(f"[mark_damaged] token={token!r}")
        except Exception:
            pass

        # try accession id
        try:
            aid = int(token)
        except Exception:
            aid = None

        if aid is not None:
            cur.execute("SELECT accession_id, status, quantity FROM devices_full WHERE accession_id = %s", (aid,))
            row = cur.fetchone()
            if not row:
                try:
                    print(f"[mark_damaged] accession_id {aid} not found")
                except Exception:
                    pass
                continue

            old_status = row.get('status') if isinstance(row, dict) else row[1]
            old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
            try:
                cur_qty = int(old_qty) if old_qty is not None else None
            except Exception:
                cur_qty = None

            if _is_token_referenced_elsewhere(cur, token, exclude_pcid):
                try:
                    print(f"[mark_damaged] accession_id={aid} still referenced elsewhere; skipping")
                except Exception:
                    pass
                continue

            if cur_qty is None:
                # single device -> mark Damaged
                if old_status != 'Damaged':
                    cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Damaged', aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Damaged', performed_by)
                    except Exception:
                        pass
                    try:
                        print(f"[mark_damaged] accession_id={aid} {old_status!r} -> 'Damaged'")
                    except Exception:
                        pass
            else:
                # stock item -> increment quantity but do not change status for the row
                new_qty = cur_qty + 1
                cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                try:
                    log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                except Exception:
                    pass
                try:
                    print(f"[mark_damaged] accession_id={aid} quantity {cur_qty} -> {new_qty} (marked as returned damaged)")
                except Exception:
                    pass
            continue

        # serial / municipal serial matches
        cur.execute("SELECT accession_id, status, quantity FROM devices_full WHERE serial_no = %s OR municipal_serial_no = %s", (token, token))
        rows = cur.fetchall()
        if rows:
            try:
                matched = [(r.get('accession_id') if isinstance(r, dict) else r[0]) for r in rows]
                print(f"[mark_damaged] serial match token={token!r} matched_accession_ids={matched}")
            except Exception:
                pass
            for row in rows:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
                old_status = row.get('status') if isinstance(row, dict) else row[1]
                old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
                try:
                    cur_qty = int(old_qty) if old_qty is not None else None
                except Exception:
                    cur_qty = None

                if _is_token_referenced_elsewhere(cur, token, exclude_pcid):
                    try:
                        print(f"[mark_damaged] accession_id={aid} still referenced elsewhere; skipping")
                    except Exception:
                        pass
                    continue

                if cur_qty is None:
                    if old_status != 'Damaged':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Damaged', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Damaged', performed_by)
                        except Exception:
                            pass
                        try:
                            print(f"[mark_damaged] accession_id={aid} {old_status!r} -> 'Damaged'")
                        except Exception:
                            pass
                else:
                    new_qty = cur_qty + 1
                    cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                    except Exception:
                        pass
                    try:
                        print(f"[mark_damaged] accession_id={aid} quantity {cur_qty} -> {new_qty} (marked as returned damaged)")
                    except Exception:
                        pass
            continue

        # partial name/brand matches - pick representative row
        like = f"%{token.lower()}%"
        cur.execute("SELECT accession_id, status, quantity FROM devices_full WHERE LOWER(item_name) LIKE %s OR LOWER(brand_model) LIKE %s ORDER BY accession_id LIMIT 1", (like, like))
        row = cur.fetchone()
        if row:
            try:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
                print(f"[mark_damaged] partial-name match token={token!r} matched_accession_id={aid}")
            except Exception:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
            old_status = row.get('status') if isinstance(row, dict) else row[1]
            old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
            try:
                cur_qty = int(old_qty) if old_qty is not None else None
            except Exception:
                cur_qty = None

            if _is_token_referenced_elsewhere(cur, token, exclude_pcid):
                try:
                    print(f"[mark_damaged] accession_id={aid} still referenced elsewhere; skipping")
                except Exception:
                    pass
            else:
                if cur_qty is None:
                    if old_status != 'Damaged':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Damaged', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Damaged', performed_by)
                        except Exception:
                            pass
                        try:
                            print(f"[mark_damaged] accession_id={aid} {old_status!r} -> 'Damaged'")
                        except Exception:
                            pass
                else:
                    new_qty = cur_qty + 1
                    cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                    except Exception:
                        pass
                    try:
                        print(f"[mark_damaged] accession_id={aid} quantity {cur_qty} -> {new_qty} (marked as returned damaged)")
                    except Exception:
                        pass
        continue

        # serial / municipal serial matches
        cur.execute("SELECT accession_id, status, quantity, device_type FROM devices_full WHERE serial_no = %s OR municipal_serial_no = %s", (token, token))
        rows = cur.fetchall()
        if rows:
            try:
                matched = [(r.get('accession_id') if isinstance(r, dict) else r[0]) for r in rows]
                print(f"[release] serial/municipal match token={token!r} matched_accession_ids={matched}")
            except Exception:
                pass
            for row in rows:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
                old_status = row.get('status') if isinstance(row, dict) else row[1]
                old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
                try:
                    cur_qty = int(old_qty) if old_qty is not None else None
                except Exception:
                    cur_qty = None

                if _is_token_referenced_elsewhere(cur, token, exclude_pcid):
                    try:
                        print(f"[release] accession_id={aid} still referenced elsewhere; skipping")
                    except Exception:
                        pass
                    continue

                if cur_qty is None:
                    if old_status != 'Available':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Available', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Available', performed_by)
                        except Exception:
                            pass
                        try:
                            print(f"[release] updated accession_id={aid} {old_status!r} -> 'Available'")
                        except Exception:
                            pass
                else:
                    new_qty = cur_qty + 1
                    cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                    except Exception:
                        pass
                    if old_status != 'Available':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Available', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Available', performed_by)
                        except Exception:
                            pass
                    try:
                        print(f"[release] accession_id={aid} quantity {cur_qty} -> {new_qty} and status -> 'Available'")
                    except Exception:
                        pass
            continue

        # partial name/brand matches
        like = f"%{token.lower()}%"
        # For partial-name releases pick one representative row (likely the consumable row)
        cur.execute("SELECT accession_id, status, quantity, device_type FROM devices_full WHERE LOWER(item_name) LIKE %s OR LOWER(brand_model) LIKE %s ORDER BY accession_id LIMIT 1", (like, like))
        row = cur.fetchone()
        if row:
            try:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
                print(f"[release] partial-name match token={token!r} matched_accession_id={aid}")
            except Exception:
                aid = row.get('accession_id') if isinstance(row, dict) else row[0]
            old_status = row.get('status') if isinstance(row, dict) else row[1]
            old_qty = row.get('quantity') if isinstance(row, dict) else row[2]
            try:
                cur_qty = int(old_qty) if old_qty is not None else None
            except Exception:
                cur_qty = None

            if _is_token_referenced_elsewhere(cur, token, exclude_pcid):
                try:
                    print(f"[release] accession_id={aid} still referenced elsewhere; skipping")
                except Exception:
                    pass
            else:
                if cur_qty is None:
                    if old_status != 'Available':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Available', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Available', performed_by)
                        except Exception:
                            pass
                        try:
                            print(f"[release] updated accession_id={aid} {old_status!r} -> 'Available'")
                        except Exception:
                            pass
                else:
                    new_qty = cur_qty + 1
                    cur.execute("UPDATE devices_full SET quantity = %s WHERE accession_id = %s", (new_qty, aid))
                    try:
                        log_inventory_action('DEVICE', aid, 'UPDATE', 'quantity', cur_qty, new_qty, performed_by)
                    except Exception:
                        pass
                    if old_status != 'Available':
                        cur.execute("UPDATE devices_full SET status = %s WHERE accession_id = %s", ('Available', aid))
                        try:
                            log_inventory_action('DEVICE', aid, 'UPDATE', 'status', old_status, 'Available', performed_by)
                        except Exception:
                            pass
                    try:
                        print(f"[release] accession_id={aid} quantity {cur_qty} -> {new_qty} and status -> 'Available'")
                    except Exception:
                        pass

   

@manage_pc_bp.route("/delete-pc/<int:pcid>", methods=["POST"])
def delete_pc(pcid):
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    conn = get_db_connection()
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            # Fetch only existing part columns so we can release referenced devices after archiving
            parts = _existing_pc_part_columns(cur)
            if parts:
                cols = ', '.join(parts)
                cur.execute(f"SELECT {cols} FROM pcinfofull WHERE pcid = %s", (pcid,))
                pc_row = cur.fetchone()
            else:
                pc_row = {}

            # Soft-archive the PC first so subsequent reference checks exclude it
            cur.execute("UPDATE pcinfofull SET is_archived = 1, deleted_at = NOW() WHERE pcid = %s", (pcid,))

            # Release referenced device items if they are not referenced by any other active PC
            performed_by = None
            try:
                performed_by = session.get('user', {}).get('user_id')
            except Exception:
                performed_by = None

            parts_to_check = parts
            if pc_row:
                for field in parts_to_check:
                    try:
                        _release_device_values_if_unreferenced(cur, pc_row.get(field), exclude_pcid=pcid, performed_by=performed_by)
                    except Exception:
                        pass

            conn.commit()
            print(f"PC with ID {pcid} archived (soft deleted).")
        if is_ajax:
            return jsonify(success=True, message="PC archived successfully.")
        flash("PC archived successfully.", "success")
    except Exception as e:
        conn.rollback()
        if is_ajax:
            return jsonify(success=False, error=str(e)), 500
        flash(str(e), "error")
    finally:
        conn.close()
    return redirect(url_for('manage_inventory.inventory_load'))


@manage_pc_bp.route('/filter-pcs', methods=['GET'])
def filter_pcs():
    conn = get_db_connection()

    try:
        args = request.args

        department_id = args.get('department_id')
        status = args.get('status')
        location = args.get('location')
        accountable = args.get('accountable')
        serial_no = args.get('serial_no')

        date_from = args.get('date_from')
        date_to = args.get('date_to')

        risk_level = args.get('risk_level')
        health_min = args.get('health_min', type=int)
        health_max = args.get('health_max', type=int)

        last_checked_from = args.get('last_checked_from')
        last_checked_to = args.get('last_checked_to')

        overdue_only = args.get('overdue') == '1'
        needs_checking = args.get('needs_checking') == '1'

        search = args.get('search')

        # Determine whether the table has `is_archived` to avoid SQL errors on older schemas
        try:
            with conn.cursor() as _cur:
                _cur.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'pcinfofull' AND COLUMN_NAME = 'is_archived'")
                _row = _cur.fetchone()
                has_is_arch = bool(_row and (int(_row.get('cnt', 0)) if isinstance(_row, dict) else _row[0] > 0))
        except Exception:
            has_is_arch = False

        # Build list of existing part columns and construct the SELECT dynamically
        try:
            with conn.cursor() as _cur:
                parts = _existing_pc_part_columns(_cur)
        except Exception:
            parts = []

        parts_select = ''
        if parts:
            parts_select = ',\n                ' + ',\n                '.join([f"p.{p}" for p in parts])

        query = f"""
            SELECT 
                p.pcid,
                p.pcname,
                p.department_id,
                d.department_name,
                p.location,
                p.quantity,
                p.acquisition_cost,
                p.date_acquired,
                p.accountable,
                p.serial_no,
                p.municipal_serial_no,
                p.status,
                p.note,

                -- health & maintenance
                p.health_score,
                CASE WHEN LOWER(TRIM(p.status)) IN ('damaged', 'damage', 'unusable') THEN 'High' ELSE p.risk_level END AS risk_level,
                p.last_checked,
                p.maintenance_interval_days
                {parts_select}

            FROM pcinfofull p
            LEFT JOIN departments d ON p.department_id = d.department_id
            WHERE 1=1
        """

        if has_is_arch:
            query += " AND COALESCE(p.is_archived, 0) = 0"

        params = []
        # Normalize status parameter and default to excluding surrendered PCs
        status_norm = status.strip().lower() if isinstance(status, str) and status.strip() else None
        if status_norm != 'surrendered':
            query += " AND LOWER(COALESCE(p.status, '')) != 'surrendered'"

        if department_id:
            query += " AND p.department_id = %s"
            params.append(department_id)

        if status:
            query += " AND LOWER(COALESCE(p.status, '')) = %s"
            params.append(status_norm)

        if location:
            query += " AND p.location LIKE %s"
            params.append(f"%{location}%")

        if accountable:
            query += " AND p.accountable LIKE %s"
            params.append(f"%{accountable}%")

        if serial_no:
            query += """
                AND (
                    p.serial_no LIKE %s
                    OR p.municipal_serial_no LIKE %s
                )
            """
            params.extend([f"%{serial_no}%", f"%{serial_no}%"])

        if date_from and date_to:
            query += " AND p.date_acquired BETWEEN %s AND %s"
            params.extend([date_from, date_to])

        # 🔹 NEW FILTERS

        if risk_level:
            query += " AND (CASE WHEN LOWER(TRIM(p.status)) IN ('damaged', 'damage', 'unusable') THEN 'High' ELSE p.risk_level END) = %s"
            params.append(risk_level)

        if health_min is not None:
            query += " AND p.health_score >= %s"
            params.append(health_min)

        if health_max is not None:
            query += " AND p.health_score <= %s"
            params.append(health_max)

        if last_checked_from and last_checked_to:
            query += " AND p.last_checked BETWEEN %s AND %s"
            params.extend([last_checked_from, last_checked_to])

        if overdue_only:
            query += """
                AND (
                    p.last_checked IS NULL
                    OR DATE_ADD(
                        p.last_checked,
                        INTERVAL GREATEST(
                            1,
                            CASE
                                WHEN IFNULL(p.maintenance_interval_days, 30) < 365
                                    THEN IFNULL(p.maintenance_interval_days, 30) * 365
                                ELSE IFNULL(p.maintenance_interval_days, 30)
                            END
                        ) DAY
                    ) < CURDATE()
                )
            """

        if needs_checking:
            query += " AND p.status = 'Needs Checking'"

        if search:
            search_cols = ['p.pcname'] + ([f"p.{p}" for p in parts] if parts else [])
            query += " AND (\n" + "\n OR ".join([f"{col} LIKE %s" for col in search_cols]) + "\n)\n"
            params.extend([f"%{search}%"] * len(search_cols))

        query += " ORDER BY p.pcid DESC"

        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            cur.execute(query, params)
            pcs = cur.fetchall()

        return jsonify(pcs)

    except Exception as e:
        print(f"❌ Error filtering PCs: {e}")
        return jsonify({"error": "Error filtering PCs"}), 500

    finally:
        conn.close()

@manage_pc_bp.route('/add-pcinfofull', methods=['POST'])
def add_pcinfofull():
    conn = get_db_connection()
    data = request.form.to_dict()

    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            # 🔹 Check inputs: require at least one identifier (serial_no or municipal_serial_no)
            serial = (data.get('serial_no') or '').strip() or None
            municipal = (data.get('municipal_serial_no') or '').strip() or None

            if not serial and not municipal:
                return jsonify({"success": False, "error": "Provide at least Serial No or Municipal Serial No."}), 400

            # 🔹 Conditional duplicate check depending on which identifiers were provided
            if serial and municipal:
                cur.execute("SELECT COUNT(*) AS count FROM pcinfofull WHERE serial_no = %s OR municipal_serial_no = %s", (serial, municipal))
            elif serial:
                cur.execute("SELECT COUNT(*) AS count FROM pcinfofull WHERE serial_no = %s", (serial,))
            else:
                cur.execute("SELECT COUNT(*) AS count FROM pcinfofull WHERE municipal_serial_no = %s", (municipal,))

            duplicate = cur.fetchone()['count']
            if duplicate > 0:
                return jsonify({"success": False, "error": "Duplicate entry: Serial No or Municipal Serial No already exists."}), 400

            # Default new PCs to Available (unless explicitly set) and initialize check/health values
            if not data.get('status'):
                data['status'] = 'Available'

            # Reuse surrendered/damaged/archived numbered slot first for single add.
            department_id = data.get('department_id')
            if department_id:
                # Fetch department info
                cur.execute(
                    "SELECT department_name, department_code, COALESCE(max_pc_allowed, 0) AS max_pc_allowed FROM departments WHERE department_id = %s",
                    (department_id,)
                )
                dept_row = cur.fetchone()

                if dept_row:
                    dept_code = (dept_row.get('department_code') or dept_row.get('department_name') or '').strip().lower().replace(' ', '-')
                    base_pcname = f"pc-{dept_code}"

                    # Check if pcinfofull has is_archived column
                    try:
                        cur.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'pcinfofull' AND COLUMN_NAME = 'is_archived'")
                        _row = cur.fetchone()
                        has_is_arch = bool(_row and (int(_row.get('cnt', 0)) if isinstance(_row, dict) else _row[0] > 0))
                    except Exception:
                        has_is_arch = False

                    # Compute current active PC total (exclude surrendered, damaged/unusable, and archived if column exists)
                    # Compute current active PC total (exclude surrendered, damaged/unusable, and archived if column exists)
                    pc_active_arch_cond = 'AND COALESCE(p.is_archived, 0) = 0' if has_is_arch else ''
                    cur.execute(f"""
                        SELECT COUNT(DISTINCT p.pcid) AS active_total
                        FROM pcinfofull p
                        WHERE p.department_id = %s {pc_active_arch_cond}
                          AND LOWER(COALESCE(p.status, '')) NOT IN ('surrendered', 'damaged', 'damage', 'unusable')
                    """, (department_id,))
                    active_row = cur.fetchone()
                    active_total = int(active_row.get('active_total', 0) if active_row else 0)

                    # Respect max_pc_allowed (0 = no limit)
                    try:
                        max_pc_allowed = int(dept_row.get('max_pc_allowed') or 0)
                    except Exception:
                        max_pc_allowed = 0

                    # Single add counts as one PC
                    new_quantity = 1

                    if max_pc_allowed > 0 and (active_total + new_quantity) > max_pc_allowed:
                        return jsonify({
                            "success": False,
                            "error": f"Cannot add PC: department has reached its PC limit ({max_pc_allowed}). Active PCs: {active_total}."
                        }), 400

                    # Get existing PC names/status for that department (include is_archived if available)
                    if has_is_arch:
                        cur.execute("""
                            SELECT pcname, status, COALESCE(is_archived, 0) AS is_archived
                            FROM pcinfofull
                            WHERE department_id = %s AND pcname LIKE %s
                        """, (department_id, f"{base_pcname}-%"))
                    else:
                        cur.execute("""
                            SELECT pcname, status, 0 AS is_archived
                            FROM pcinfofull
                            WHERE department_id = %s AND pcname LIKE %s
                        """, (department_id, f"{base_pcname}-%"))
                    existing_rows = cur.fetchall()

                    pattern = re.compile(rf"^{re.escape(base_pcname)}-(\d+)$", re.IGNORECASE)
                    active_numbers = set()
                    available_numbers = set()

                    for row in existing_rows:
                        pc_name = (row.get('pcname') or '').strip()
                        match = pattern.match(pc_name)
                        if not match:
                            continue

                        seq_number = int(match.group(1))
                        row_status = str(row.get('status') or '').strip().lower()
                        row_archived = int(row.get('is_archived') or 0)

                        # Treat surrendered, damaged/unusable, or archived rows as available for reuse
                        if row_archived == 1 or row_status in ('surrendered', 'damaged', 'damage', 'unusable'):
                            available_numbers.add(seq_number)
                        else:
                            active_numbers.add(seq_number)

                    reusable_numbers = sorted(num for num in available_numbers if num not in active_numbers)
                    if reusable_numbers:
                        data['pcname'] = f"{base_pcname}-{reusable_numbers[0]:02d}"

            try:
                dept_assigned = bool(department_id)
            except Exception:
                dept_assigned = False

            # Do not override explicit status choices from the form.


            # 🔹 Insert new PC (only include existing part columns)
            parts = _existing_pc_part_columns(cur)
            base_cols = [
                'pcname', 'department_id', 'location', 'quantity', 'acquisition_cost',
                'date_acquired', 'accountable', 'serial_no', 'municipal_serial_no', 'status',
                'note', 'maintenance_interval_days'
            ]
            cols = base_cols + parts + ['last_checked', 'health_score', 'risk_level']
            placeholders = ','.join(['%s'] * (len(base_cols) + len(parts))) + ", CURDATE(), 100, 'Low'"
            sql = f"INSERT INTO pcinfofull ({', '.join(cols)}) VALUES ({placeholders})"
            params = [
                data.get('pcname'), data.get('department_id'), data.get('location'), data.get('quantity'),
                data.get('acquisition_cost'), data.get('date_acquired'), data.get('accountable'),
                serial, municipal, data.get('status'), data.get('note'), data.get('maintenance_interval_days')
            ] + [data.get(p) for p in parts]

            cur.execute(sql, params)

            # Ensure DB defaults/triggers can't leave the newly created PC in an incorrect state
            new_pcid = cur.lastrowid

            # Decide final status for the newly inserted PC (respect explicit choices)
            status_to_set = data.get('status')
            if not status_to_set or str(status_to_set).strip() == '':
                status_to_set = 'IN USE' if dept_assigned else 'Available'

            # Persist canonical defaults for last_checked/health/risk and chosen status
            cur.execute("""
                UPDATE pcinfofull
                SET
                    status = %s,
                    last_checked = CURDATE(),
                    health_score = 100,
                    risk_level = 'Low'
                WHERE pcid = %s
            """, (status_to_set, new_pcid))
            # Reflect canonical status back into data for downstream logic
            data['status'] = status_to_set
            # After committing the new PC, mark any referenced device items as IN USE
            performed_by = None
            try:
                performed_by = session.get('user', {}).get('user_id')
            except Exception:
                performed_by = None

            parts_to_check = parts
            for part_field in parts_to_check:
                _mark_device_values_in_use(cur, data.get(part_field), performed_by)

            conn.commit()

        return jsonify({"success": True, "message": "PC added successfully!"})

    except Exception as e:
        conn.rollback()
        print(f"❌ Error adding PC: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        conn.close()
    manage_inventory.risk




@manage_pc_bp.route('/update-pcinfofull', methods=['POST'])
def update_pcinfofull():
    if 'user' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    conn = get_db_connection()
    data = request.form
    user_id = session['user']['user_id']
    # Ensure status_param is defined early to avoid UnboundLocalError
    status_param = (data.get('status') or '').strip()

    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:

            # 🔹 1. Duplicate check (require at least one identifier)
            serial = (data.get('serial_no') or '').strip() or None
            municipal = (data.get('municipal_serial_no') or '').strip() or None
            pcid = data.get('pcid')

            # Debug: log incoming identifiers and requested status
            try:
                print(f"[update_pc] incoming pcid={pcid!r} status_raw={repr(data.get('status'))} department_id={repr(data.get('department_id'))}")
            except Exception:
                pass

            if not serial and not municipal:
                return jsonify({"success": False, "error": "Provide at least Serial No or Municipal Serial No."}), 400

            if serial and municipal:
                cur.execute("SELECT COUNT(*) AS count FROM pcinfofull WHERE (serial_no = %s OR municipal_serial_no = %s) AND pcid != %s", (serial, municipal, pcid))
            elif serial:
                cur.execute("SELECT COUNT(*) AS count FROM pcinfofull WHERE serial_no = %s AND pcid != %s", (serial, pcid))
            else:
                cur.execute("SELECT COUNT(*) AS count FROM pcinfofull WHERE municipal_serial_no = %s AND pcid != %s", (municipal, pcid))

            if cur.fetchone()['count'] > 0:
                return jsonify({"success": False, "error": "Duplicate entry: Serial No or Municipal Serial No already exists."}), 400

            # 🔹 2. Fetch OLD values
            cur.execute("SELECT * FROM pcinfofull WHERE pcid = %s", (data['pcid'],))
            old_pc = cur.fetchone()

            if not old_pc:
                return jsonify({"success": False, "error": "PC not found"}), 404

            # 🔹 3. Perform UPDATE
            # Determine effective status: only derive when status is not provided.
            status_param = (data.get('status') or '').strip()
            dept_param = data.get('department_id')
            try:
                print(f"[update_pc] before derive status_raw={repr(data.get('status'))} dept_param={repr(dept_param)}")
            except Exception:
                pass
            if not status_param:
                status_param = 'IN USE' if dept_param else 'Available'
            try:
                print(f"[update_pc] derived status_param={repr(status_param)}")
            except Exception:
                pass

            # Build dynamic update including only existing part columns
            parts = _existing_pc_part_columns(cur)
            update_fields = [
                "pcname=%s",
                "department_id=%s",
                "location=%s",
                "quantity=%s",
                "acquisition_cost=%s",
                "date_acquired=%s",
                "accountable=%s",
                "serial_no=%s",
                "municipal_serial_no=%s",
                "status=%s",
                "risk_level=CASE WHEN LOWER(TRIM(%s)) IN ('damaged', 'damage', 'unusable') THEN 'High' ELSE risk_level END",
                "note=%s",
                "maintenance_interval_days=%s",
            ]
            params = [
                data.get('pcname'),
                data.get('department_id'),
                data.get('location'),
                data.get('quantity'),
                data.get('acquisition_cost'),
                data.get('date_acquired'),
                data.get('accountable'),
                serial,
                municipal,
                status_param,
                status_param,
                data.get('note'),
                data.get('maintenance_interval_days')
            ]

            for p in parts:
                update_fields.append(f"{p}=%s")
                params.append(data.get(p))

            params.append(data.get('pcid'))
            sql = "UPDATE pcinfofull SET\n                " + ",\n                ".join(update_fields) + "\n            WHERE pcid=%s"
            cur.execute(sql, params)

            # 🔹 4. AUDIT LOGGING
            tracked_fields = [
                'pcname', 'department_id', 'location', 'quantity',
                'acquisition_cost', 'date_acquired', 'accountable',
                'serial_no', 'municipal_serial_no', 'status', 'note',
                'maintenance_interval_days'
            ] + parts

            for field in tracked_fields:
                old_value = old_pc.get(field)
                # Use derived status_param for accurate audit if status was inferred
                if field == 'status':
                    new_value = status_param
                else:
                    new_value = data.get(field)

                # Normalize dates
                if isinstance(old_value, date):
                    old_value = old_value.strftime('%Y-%m-%d')

                # Skip if both empty
                if (old_value is None or old_value == '') and (new_value is None or new_value == ''):
                    continue

                if str(old_value) != str(new_value):
                    log_inventory_action(
                        entity_type='PC',
                        entity_id=data['pcid'],
                        action='UPDATE',
                        field_name=field,
                        old_value=old_value,
                        new_value=new_value,
                        performed_by=user_id
                    )

            # Release any previously assigned parts that are no longer present, then mark new ones as IN USE
            parts_to_check = parts

            # Build token sets
            old_tokens = set()
            for f in parts_to_check:
                try:
                    old_tokens.update(_parse_tokens(old_pc.get(f)))
                except Exception:
                    pass

            new_tokens = set()
            for f in parts_to_check:
                try:
                    new_tokens.update(_parse_tokens(data.get(f)))
                except Exception:
                    pass

            # Tokens removed in the update -> consider releasing them
            tokens_to_release = old_tokens - new_tokens
            for token in tokens_to_release:
                try:
                    _release_device_values_if_unreferenced(cur, token, exclude_pcid=pcid, performed_by=user_id)
                except Exception:
                    pass

            # Mark new/assigned tokens as IN USE
            for part_field in parts_to_check:
                try:
                    _mark_device_values_in_use(cur, data.get(part_field), user_id)
                except Exception:
                    pass

            # After performing updates, fetch current DB status for logging/verification
            try:
                cur.execute("SELECT status FROM pcinfofull WHERE pcid = %s", (data.get('pcid'),))
                _new = cur.fetchone()
                try:
                    print(f"[update_pc] DB status after update: {(_new.get('status') if _new else None)!r}")
                except Exception:
                    pass
            except Exception:
                pass

            conn.commit()

        return jsonify({"success": True, "message": "PC updated successfully!"})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        conn.close()
@manage_pc_bp.route('/replace-pc-part', methods=['POST'])
def replace_pc_part():
    if 'user' not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    # Accept JSON or form data; support single-part legacy or multiple replacements
    data = request.get_json(silent=True) or request.form
    pcid = data.get('pcid')
    replacements = {}

    # If JSON provided with 'replacements' map
    if isinstance(data, dict) and data.get('replacements'):
        replacements = data.get('replacements') or {}
    else:
        # Fall back to single-part legacy parameters
        part = data.get('part')
        new_value = data.get('new_value') if data.get('new_value') is not None else ''
        if part:
            replacements = {part: new_value}

    allowed_parts = ['monitor', 'motherboard', 'ram', 'storage', 'gpu', 'psu', 'casing', 'mouse', 'keyboard', 'other_parts']
    if not pcid or not replacements:
        return jsonify({"success": False, "error": "Invalid request parameters"}), 400

    # Validate replacement keys
    for p in list(replacements.keys()):
        if p not in allowed_parts:
            return jsonify({"success": False, "error": f"Invalid part: {p}"}), 400

    conn = get_db_connection()
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            # Determine which part columns actually exist in the DB for this install
            existing_parts = _existing_pc_part_columns(cur)

            # Determine requested parts that are not present in the DB schema.
            # Attempt to create missing columns automatically (only for
            # allowed part names) so replacements persist in dedicated cols.
            missing_parts = [p for p in list(replacements.keys()) if p not in existing_parts]
            if missing_parts:
                for p in list(missing_parts):
                    # Only add columns for known/allowed part names
                    if p not in allowed_parts:
                        continue
                    try:
                        cur.execute(f"ALTER TABLE pcinfofull ADD COLUMN `{p}` TEXT")
                        try:
                            print(f"[replace_pc_part] Added missing column `{p}` to pcinfofull")
                        except Exception:
                            pass
                    except Exception as e:
                        try:
                            print(f"[replace_pc_part] Could not add column `{p}`: {e}")
                        except Exception:
                            pass

                # Refresh list of existing parts after attempted ALTERs
                existing_parts = _existing_pc_part_columns(cur)

            # Recompute missing after potential schema changes; if still
            # missing and there's no `other_parts` column to fall back to,
            # return an error for the first unsupported part.
            missing_parts = [p for p in list(replacements.keys()) if p not in existing_parts]
            if missing_parts and 'other_parts' not in existing_parts:
                return jsonify({"success": False, "error": f"Part not supported by current schema: {missing_parts[0]}"}), 400

            # Fetch current values for only the existing parts
            if existing_parts:
                cols = ', '.join(existing_parts)
                cur.execute(f"SELECT {cols} FROM pcinfofull WHERE pcid = %s", (pcid,))
                row = cur.fetchone()
            else:
                row = {}

            if not row:
                return jsonify({"success": False, "error": "PC not found"}), 404

            performed_by = None
            try:
                performed_by = session.get('user', {}).get('user_id')
            except Exception:
                performed_by = None

            # If the caller requested replacements for part columns that do
            # not exist, merge their new-values into `other_parts` (if present)
            if missing_parts:
                tokens_to_add = []
                for p in missing_parts:
                    v = replacements.pop(p, '')
                    if v:
                        tokens_to_add.append(str(v).strip())

                base_other = ''
                try:
                    base_other = (replacements.get('other_parts') or row.get('other_parts') or '')
                    base_other = str(base_other).strip()
                except Exception:
                    base_other = ''

                if tokens_to_add:
                    if base_other:
                        combined = base_other + ', ' + ', '.join(tokens_to_add)
                    else:
                        combined = ', '.join(tokens_to_add)
                    replacements['other_parts'] = combined

            # Build sets of tokens to release: only release tokens that are
            # present in the old values but not present in the new values.
            parts_to_check = [p for p in list(replacements.keys()) if p in existing_parts]
            old_tokens = set()
            for f in parts_to_check:
                try:
                    old_tokens.update(_parse_tokens(row.get(f)))
                except Exception:
                    pass

            new_tokens = set()
            for f in parts_to_check:
                try:
                    new_tokens.update(_parse_tokens(replacements.get(f)))
                except Exception:
                    pass

            tokens_to_release = old_tokens - new_tokens
            for token in tokens_to_release:
                try:
                    # When replacing a part, the removed/old part should be marked
                    # as Damaged and returned to the Manage Item list instead of
                    # being marked Available. Use specialized helper.
                    _mark_device_values_damaged(cur, token, exclude_pcid=pcid, performed_by=performed_by)
                except Exception:
                    pass

            # Build UPDATE statement for all replacements
            set_clauses = []
            params = []
            for part, new_value in replacements.items():
                # Skip if new value equals current value to avoid no-op
                old_value = row.get(part)
                if (old_value is None and (new_value is None or new_value == '')) or str(old_value) == str(new_value):
                    continue
                set_clauses.append(f"{part} = %s")
                params.append(new_value)

            if set_clauses:
                sql = "UPDATE pcinfofull SET " + ", ".join(set_clauses) + " WHERE pcid = %s"
                params.append(pcid)
                cur.execute(sql, params)

            # Audit and mark new values as IN USE
            for part, new_value in replacements.items():
                old_value = row.get(part)
                try:
                    if str(old_value) != str(new_value):
                        try:
                            log_inventory_action('PC', pcid, 'UPDATE', part, old_value, new_value, performed_by)
                        except Exception:
                            pass
                except Exception:
                    pass

                try:
                    _mark_device_values_in_use(cur, new_value, performed_by)
                except Exception:
                    pass

            conn.commit()
        return jsonify({"success": True, "message": "Part(s) replaced successfully."})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        conn.close()


@manage_pc_bp.route('/batch_add_pcinfofull', methods=['POST'])
def batch_add_pcinfofull():
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'success': False, 'error': 'Invalid data format'}), 400

    conn = get_db_connection()
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            inserted = 0

            # 🔹 Get department info from first entry
            first_pc = data[0]
            department_id = first_pc.get('department_id')
            if not department_id:
                return jsonify({'success': False, 'error': 'Missing department ID'}), 400

            # 🔹 Get department_code and max_pc_allowed
            cur.execute("SELECT department_name, department_code, COALESCE(max_pc_allowed, 0) AS max_pc_allowed FROM departments WHERE department_id = %s", (department_id,))
            dept_row = cur.fetchone()
            if not dept_row:
                return jsonify({'success': False, 'error': 'Department not found'}), 404

            dept_code = (dept_row.get('department_code') or dept_row.get('department_name')).strip().lower().replace(' ', '-')
            base_pcname = f"pc-{dept_code}"

            # Check if pcinfofull has is_archived column
            try:
                cur.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'pcinfofull' AND COLUMN_NAME = 'is_archived'")
                _row = cur.fetchone()
                has_is_arch = bool(_row and (int(_row.get('cnt', 0)) if isinstance(_row, dict) else _row[0] > 0))
            except Exception:
                has_is_arch = False

            # Compute current active PC total (exclude surrendered, damaged/unusable, and archived if column exists)
            pc_active_arch_cond = 'AND COALESCE(p.is_archived, 0) = 0' if has_is_arch else ''
            cur.execute(f"""
                SELECT COUNT(DISTINCT p.pcid) AS active_total
                FROM pcinfofull p
                WHERE p.department_id = %s {pc_active_arch_cond}
                  AND LOWER(COALESCE(p.status, '')) NOT IN ('surrendered', 'damaged', 'damage', 'unusable')
            """, (department_id,))
            active_row = cur.fetchone()
            active_total = int(active_row.get('active_total', 0) if active_row else 0)

            # Determine total number of new PCs (skip invalid rows and duplicates)
            total_new = 0
            for pc in data:
                serial = pc.get('serial_no') or None
                municipal = pc.get('municipal_serial_no') or None
                if not serial and not municipal:
                    continue
                cur.execute(
                    """
                    SELECT COUNT(*) AS count
                    FROM pcinfofull
                    WHERE serial_no = %s OR municipal_serial_no = %s
                    """, (serial, municipal)
                )
                if cur.fetchone()['count'] > 0:
                    # duplicate/skipped
                    continue
                total_new += 1

            try:
                max_pc_allowed = int(dept_row.get('max_pc_allowed') or 0)
            except Exception:
                max_pc_allowed = 0

            if max_pc_allowed > 0 and (active_total + total_new) > max_pc_allowed:
                return jsonify({'success': False, 'error': f'Cannot add PCs: would exceed department PC limit ({max_pc_allowed}). Active: {active_total}, adding: {total_new}'}), 400

            # 🔹 Get existing PC names/status for that department (include is_archived if available)
            if has_is_arch:
                cur.execute("""
                    SELECT pcname, status, COALESCE(is_archived, 0) AS is_archived FROM pcinfofull
                    WHERE department_id = %s AND pcname LIKE %s
                """, (department_id, f"{base_pcname}-%"))
            else:
                cur.execute("""
                    SELECT pcname, status, 0 AS is_archived FROM pcinfofull
                    WHERE department_id = %s AND pcname LIKE %s
                """, (department_id, f"{base_pcname}-%"))
            existing_rows = cur.fetchall()

            # Reuse surrendered/damaged/archived slots first (e.g., if ...-21 is surrendered, next add gets ...-21).
            pattern = re.compile(rf"^{re.escape(base_pcname)}-(\d+)$", re.IGNORECASE)
            active_numbers = set()
            available_numbers = set()

            for row in existing_rows:
                pc_name = (row.get('pcname') or '').strip()
                match = pattern.match(pc_name)
                if not match:
                    continue

                seq_number = int(match.group(1))
                row_status = str(row.get('status') or '').strip().lower()
                row_archived = int(row.get('is_archived') or 0)

                if row_archived == 1 or row_status in ('surrendered', 'damaged', 'damage', 'unusable'):
                    available_numbers.add(seq_number)
                else:
                    active_numbers.add(seq_number)

            reusable_numbers = sorted(num for num in available_numbers if num not in active_numbers)
            all_seen_numbers = active_numbers.union(available_numbers)
            next_number = (max(all_seen_numbers) + 1) if all_seen_numbers else 1


            # 🔹 Pre-validate batch: require at least one identifier per row and detect in-file duplicates
            missing_rows = []
            serial_map = defaultdict(list)
            municipal_map = defaultdict(list)

            for idx, pc in enumerate(data):
                row_no = idx + 1
                s = (pc.get('serial_no') or '').strip() or None
                m = (pc.get('municipal_serial_no') or '').strip() or None
                if not s and not m:
                    missing_rows.append(row_no)
                if s:
                    serial_map[s].append(row_no)
                if m:
                    municipal_map[m].append(row_no)

            duplicate_serials = {s: rows for s, rows in serial_map.items() if len(rows) > 1}
            duplicate_municipals = {m: rows for m, rows in municipal_map.items() if len(rows) > 1}

            if missing_rows or duplicate_serials or duplicate_municipals:
                issues = []
                if missing_rows:
                    issues.append(f"Rows missing both Serial No and Municipal Serial No: {', '.join(map(str, missing_rows))}")
                if duplicate_serials:
                    items = [f"'{s}' (rows {', '.join(map(str, rows))})" for s, rows in duplicate_serials.items()]
                    issues.append(f"Duplicate Serial No in batch: {', '.join(items)}")
                if duplicate_municipals:
                    items = [f"'{m}' (rows {', '.join(map(str, rows))})" for m, rows in duplicate_municipals.items()]
                    issues.append(f"Duplicate Municipal Serial No in batch: {', '.join(items)}")

                return jsonify({'success': False, 'error': '; '.join(issues)}), 400

            # 🔹 Check database for existing identifiers referenced in batch
            db_conflicts = []
            if serial_map:
                placeholders = ','.join(['%s'] * len(serial_map))
                cur.execute(f"SELECT serial_no FROM pcinfofull WHERE serial_no IN ({placeholders})", list(serial_map.keys()))
                for r in cur.fetchall():
                    db_conflicts.append(f"Serial '{r.get('serial_no')}'")
            if municipal_map:
                placeholders = ','.join(['%s'] * len(municipal_map))
                cur.execute(f"SELECT municipal_serial_no FROM pcinfofull WHERE municipal_serial_no IN ({placeholders})", list(municipal_map.keys()))
                for r in cur.fetchall():
                    db_conflicts.append(f"Municipal '{r.get('municipal_serial_no')}'")

            if db_conflicts:
                return jsonify({'success': False, 'error': 'Duplicate in database: ' + ', '.join(db_conflicts)}), 400

            # 🔹 Insert each PC
            for pc in data:
                if not pc.get('status'):
                    pc['status'] = 'Available'

                if reusable_numbers:
                    assigned_number = reusable_numbers.pop(0)
                else:
                    while next_number in active_numbers:
                        next_number += 1
                    assigned_number = next_number
                    next_number += 1

                active_numbers.add(assigned_number)
                pc['pcname'] = f"{base_pcname}-{assigned_number:02d}"

                # Insert using only existing part columns
                parts = _existing_pc_part_columns(cur)
                base_cols = [
                    'pcname', 'department_id', 'location', 'quantity', 'acquisition_cost',
                    'date_acquired', 'accountable', 'serial_no', 'municipal_serial_no', 'status',
                    'note', 'maintenance_interval_days'
                ]
                cols = base_cols + parts + ['last_checked', 'health_score', 'risk_level', 'created_at', 'updated_at']
                placeholders = ','.join(['%s'] * (len(base_cols) + len(parts))) + ", CURDATE(), 100, 'Low', NOW(), NOW()"
                sql = f"INSERT INTO pcinfofull ({', '.join(cols)}) VALUES ({placeholders})"
                params = [pc.get(c) for c in base_cols] + [pc.get(p) for p in parts]
                cur.execute(sql, params)
                # mark referenced device items as IN USE for this pc
                performed_by = None
                try:
                    performed_by = session.get('user', {}).get('user_id')
                except Exception:
                    performed_by = None

                for part_field in parts:
                    _mark_device_values_in_use(cur, pc.get(part_field), performed_by)

                inserted += 1

        conn.commit()
        return jsonify({
            'success': True,
        #     'message': f'Successfully added {inserted} PCs under {dept_row['department_name']} ({dept_code.upper()})'
        # 
        })

    except Exception as e:
        print("❌ Batch insert error:", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@manage_pc_bp.route('/manage_pc/export-selected-pcs', methods=['POST'])
def export_selected_pcs():

    data = request.get_json(silent=True) or {}
    pcids = data.get("pcids") or []

    try:
        pcids = [int(pid) for pid in pcids]
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid PC selection"}), 400

    if not pcids:
        return jsonify({"error": "No PCs selected"}), 400

    conn = get_db_connection()

    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:

            format_strings = ','.join(['%s'] * len(pcids))

            cur.execute(f"""
                SELECT
                    p.pcid,
                    p.pcname,
                    COALESCE(d.department_name, '') AS department,
                    p.location,
                    p.acquisition_cost,
                    p.date_acquired,
                    p.accountable,
                    p.serial_no,
                    p.municipal_serial_no,
                    p.status
                FROM pcinfofull p
                LEFT JOIN departments d ON p.department_id = d.department_id
                WHERE p.pcid IN ({format_strings})
                ORDER BY p.pcid
            """, pcids)

            rows = cur.fetchall()

        if not rows:
            return jsonify({"error": "No matching PCs found for export"}), 404

        df = pd.DataFrame(rows).rename(columns={
            "pcid": "PC ID",
            "pcname": "PC Name",
            "department": "Department",
            "location": "Location",
            "acquisition_cost": "Acquisition Cost",
            "date_acquired": "Date Acquired",
            "accountable": "Accountable",
            "serial_no": "Serial No.",
            "municipal_serial_no": "Municipal Serial No.",
            "status": "Status",
        })

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = 'Selected PCs'
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]

            header_fill = PatternFill(fill_type='solid', fgColor='4F46E5')
            header_font = Font(color='FFFFFF', bold=True)
            thin_side = Side(style='thin', color='D1D5DB')
            cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            for col_idx, col_name in enumerate(df.columns, start=1):
                header_cell = ws.cell(row=1, column=col_idx)
                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.border = cell_border

                values = [str(col_name)]
                for value in df.iloc[:, col_idx - 1].tolist():
                    values.append('' if value is None else str(value))
                max_len = max(len(v) for v in values)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 36)

            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = cell_border

            column_map = {name: idx + 1 for idx, name in enumerate(df.columns)}

            if 'Acquisition Cost' in column_map:
                col = column_map['Acquisition Cost']
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col).number_format = '#,##0.00'

            if 'Date Acquired' in column_map:
                col = column_map['Date Acquired']
                for row_idx in range(2, ws.max_row + 1):
                    date_cell = ws.cell(row=row_idx, column=col)
                    date_cell.number_format = 'yyyy-mm-dd'
                    date_cell.alignment = Alignment(horizontal='center', vertical='center')

            for name in ('PC ID', 'Status'):
                if name in column_map:
                    col = column_map[name]
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="selected_pcs.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("❌ Export error:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        conn.close()

@manage_pc_bp.route('/manage_pc/import-pcs-excel', methods=['POST'])
def import_pcs_excel():
    file = request.files.get("file")
    duplicate_option = request.form.get("duplicate_option", "skip")

    if not file:
        return jsonify({"success": False, "error": "No file uploaded"})

    if duplicate_option not in {"skip", "ignore", "overwrite"}:
        duplicate_option = "skip"

    conn = None

    try:
        df = pd.read_excel(file)

        # Normalize incoming headers so both template exports and raw imports are supported.
        def normalize_header(name):
            text = str(name).strip().lower()
            text = re.sub(r"[^a-z0-9]+", "_", text)
            return text.strip("_")

        df = df.rename(columns={col: normalize_header(col) for col in df.columns})

        alias_map = {
            "pc_name": "pcname",
            "department_name": "department",
        }
        for old_name, new_name in alias_map.items():
            if old_name in df.columns and new_name not in df.columns:
                df[new_name] = df[old_name]

        # Lightweight cleaning helper used for pre-validation and per-row processing
        def clean_value(value):
            if pd.isna(value):
                return None
            if isinstance(value, str):
                value = value.strip()
                return value or None
            if isinstance(value, pd.Timestamp):
                return value.date()
            return value

        # Pre-validate uploaded file for duplicate serial / municipal numbers
        serial_rows = defaultdict(list)
        municipal_rows = defaultdict(list)
        missing_both_rows = []

        for idx, row in df.iterrows():
            row_number = int(idx) + 2  # Excel row (header at row 1)
            serial = clean_value(row.get("serial_no"))
            municipal = clean_value(row.get("municipal_serial_no"))

            if not serial and not municipal:
                missing_both_rows.append(row_number)

            if serial:
                serial_rows[serial].append(row_number)

            if municipal:
                municipal_rows[municipal].append(row_number)

        duplicate_serials = {s: rows for s, rows in serial_rows.items() if len(rows) > 1}
        duplicate_municipals = {m: rows for m, rows in municipal_rows.items() if len(rows) > 1}

        # If the uploaded file contains duplicates within itself, fail fast and return details
        if duplicate_serials or duplicate_municipals or missing_both_rows:
            issues = []
            if duplicate_serials:
                items = [f"'{s}' (rows {', '.join(map(str, rows))})" for s, rows in duplicate_serials.items()]
                issues.append(f"Duplicate Serial No in file: {', '.join(items)}")
            if duplicate_municipals:
                items = [f"'{m}' (rows {', '.join(map(str, rows))})" for m, rows in duplicate_municipals.items()]
                issues.append(f"Duplicate Municipal Serial No in file: {', '.join(items)}")
            if missing_both_rows:
                issues.append(f"Rows missing both Serial No and Municipal Serial No: {', '.join(map(str, missing_both_rows))}")

            return jsonify({
                "success": False,
                "error": "; ".join(issues)
            }), 400

        conn = get_db_connection()
        cur = conn.cursor(pymysql.cursors.DictCursor)

        added = 0
        updated = 0
        skipped = 0

        department_cache = {}

        for _, row in df.iterrows():

            pcname = clean_value(row.get("pcname"))
            serial = clean_value(row.get("serial_no"))
            municipal = clean_value(row.get("municipal_serial_no"))
            location = clean_value(row.get("location"))
            acquisition_cost = clean_value(row.get("acquisition_cost"))
            date_acquired = clean_value(row.get("date_acquired"))
            accountable = clean_value(row.get("accountable"))
            status = clean_value(row.get("status")) or "Available"
            status_text = str(status).strip()
            normalized_status_text = status_text.lower()
            imported_risk = "High" if normalized_status_text in {"damaged", "damage", "unusable"} else "Low"

            department_id = clean_value(row.get("department_id"))
            department_name = clean_value(row.get("department"))

            if not department_id and department_name:
                if department_name not in department_cache:
                    cur.execute(
                        "SELECT department_id FROM departments WHERE department_name = %s LIMIT 1",
                        (department_name,)
                    )
                    dept_row = cur.fetchone()
                    department_cache[department_name] = dept_row["department_id"] if dept_row else None
                department_id = department_cache.get(department_name)

            if not serial and not municipal:
                skipped += 1
                continue

            existing = None
            if serial and municipal:
                cur.execute(
                    "SELECT pcid FROM pcinfofull WHERE serial_no = %s OR municipal_serial_no = %s LIMIT 1",
                    (serial, municipal)
                )
                existing = cur.fetchone()
            elif serial:
                cur.execute(
                    "SELECT pcid FROM pcinfofull WHERE serial_no = %s LIMIT 1",
                    (serial,)
                )
                existing = cur.fetchone()
            elif municipal:
                cur.execute(
                    "SELECT pcid FROM pcinfofull WHERE municipal_serial_no = %s LIMIT 1",
                    (municipal,)
                )
                existing = cur.fetchone()

            if existing:
                if duplicate_option == "skip":
                    skipped += 1
                    continue

                if duplicate_option == "ignore":
                    skipped += 1
                    continue

                if duplicate_option == "overwrite":
                    cur.execute("""
                        UPDATE pcinfofull SET
                        pcname = COALESCE(%s, pcname),
                        department_id = COALESCE(%s, department_id),
                        location = COALESCE(%s, location),
                        acquisition_cost = COALESCE(%s, acquisition_cost),
                        date_acquired = COALESCE(%s, date_acquired),
                        accountable = COALESCE(%s, accountable),
                        status = COALESCE(%s, status),
                        risk_level = CASE WHEN LOWER(TRIM(COALESCE(%s, status))) IN ('damaged', 'damage', 'unusable') THEN 'High' ELSE risk_level END
                        WHERE pcid=%s
                    """, (
                        pcname,
                        department_id,
                        location,
                        acquisition_cost,
                        date_acquired,
                        accountable,
                        status,
                        status,
                        existing["pcid"]
                    ))

                    updated += 1
                    continue

            if not pcname:
                skipped += 1
                continue

            cur.execute("""
                INSERT INTO pcinfofull
                (pcname,department_id,location,acquisition_cost,date_acquired,
                 accountable,serial_no,municipal_serial_no,status,
                 last_checked,health_score,risk_level)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s, CURDATE(), 100, %s)
            """, (
                pcname,
                department_id,
                location,
                acquisition_cost,
                date_acquired,
                accountable,
                serial,
                municipal,
                status,
                imported_risk
            ))

            added += 1

        conn.commit()

        return jsonify({
            "success": True,
            "added": added,
            "updated": updated,
            "skipped": skipped
        })

    except Exception as e:
        print("Import error:", e)
        return jsonify({"success": False, "error": str(e)})

    finally:
        if conn:
            conn.close()