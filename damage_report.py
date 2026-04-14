from flask import Blueprint, jsonify, request, render_template, send_file
import pymysql
from db import get_db_connection
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from utils.permissions import check_permission

damage_report_bp = Blueprint('damage_report_bp', __name__, template_folder='templates')


# -------------------------------
# 1️⃣ Render Damage Report Page
# -------------------------------
@damage_report_bp.route('/damage-report')
@check_permission('damage_report', 'view')
def damage_report_page():
    """Render the Damage Report HTML page."""
    return render_template('damage_report.html')


@damage_report_bp.route('/get-damage-reports', methods=['GET'])
@check_permission('damage_report', 'view')
def loadDamageReports():

    conn = None
    try:
        conn = get_db_connection()

        # Filters
        name = request.args.get('name', '').strip()
        category = request.args.get('category', '').strip()
        department = request.args.get('department', '').strip()
        severity = request.args.get('severity', '').strip()  # Changed from 'status' to 'severity'
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()

        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        offset = (page - 1) * per_page

        query = """
            SELECT 
                d.accession_id AS id,
                d.item_name AS name,
                'Device' AS category,
                dept.department_name AS department,
                '' AS location,
                d.accountable,
                COALESCE(ddr.damage_type, 'Minor') AS damage_type,
                ddr.description AS damage_description,
                d.acquisition_cost,
                d.date_acquired
            FROM devices_full d
            LEFT JOIN departments dept ON dept.department_id = d.department_id
            LEFT JOIN device_damage_reports ddr ON ddr.accession_id = d.accession_id
            WHERE d.status = 'Damaged'
            UNION ALL

            SELECT 
                p.pcid AS id,
                p.pcname AS name,
                'PC' AS category,
                dept.department_name AS department,
                p.location,
                p.accountable,
                COALESCE(dr.damage_type, 'Minor') AS damage_type,
                dr.description AS damage_description,
                p.acquisition_cost,
                p.date_acquired
            FROM pcinfofull p
            LEFT JOIN departments dept ON dept.department_id = p.department_id
            LEFT JOIN damage_reports dr ON dr.pcid = p.pcid
            WHERE p.status = 'Damaged'
        """

        filters = []
        params = []

        if name:
            filters.append("name LIKE %s")
            params.append(f"%{name}%")
        if category and category.lower() != "all":
            filters.append("category = %s")
            params.append(category)
        if department and department.lower() != "all":
            filters.append("department = %s")
            params.append(department)
        if severity and severity.lower() != "all":
            filters.append("damage_type = %s")
            params.append(severity)
        if date_from:
            filters.append("date_acquired >= %s")
            params.append(date_from)
        if date_to:
            filters.append("date_acquired <= %s")
            params.append(date_to)

        if filters:
            query = f"SELECT * FROM ({query}) AS combined WHERE {' AND '.join(filters)}"

        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            # ---------- TOTAL COUNT ----------
            count_query = f"SELECT COUNT(*) as total FROM ({query}) as count_table"
            cursor.execute(count_query, params)
            total = cursor.fetchone()["total"]

            # ---------- PAGINATION ----------
            query += " ORDER BY date_acquired DESC LIMIT %s OFFSET %s"
            params_with_pagination = params + [per_page, offset]

            cursor.execute(query, params_with_pagination)
            data = cursor.fetchall()

        total_pages = (total + per_page - 1) // per_page

        return jsonify({
            "data": data,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages,
                "has_prev": page > 1,
                "has_next": page < total_pages
            }
        })
    except Exception as e:
        return jsonify({
            "data": [],
            "pagination": {
                "page": 1,
                "per_page": 10,
                "total": 0,
                "total_pages": 0,
                "has_prev": False,
                "has_next": False
            },
            "error": str(e)
        }), 500
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


@damage_report_bp.route('/export-damage-reports', methods=['GET'])
@check_permission('damage_report', 'view')
def export_damage_reports():
    """Export filtered damage reports as Excel file."""
    conn = None
    try:
        conn = get_db_connection()

        # Same filters as /get-damage-reports
        name = request.args.get('name', '').strip()
        category = request.args.get('category', '').strip()
        department = request.args.get('department', '').strip()
        severity = request.args.get('severity', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()

        query = """
            SELECT
                d.item_name AS `Item / PC Name`,
                'Device' AS Category,
                dept.department_name AS Facilities,
                '' AS Location,
                d.accountable AS Accountable,
                COALESCE(ddr.damage_type, 'Minor') AS Severity,
                ddr.description AS `Damage Description`,
                d.acquisition_cost AS `Acquisition Cost`,
                d.date_acquired AS `Date Acquired`
            FROM devices_full d
            LEFT JOIN departments dept ON dept.department_id = d.department_id
            LEFT JOIN device_damage_reports ddr ON ddr.accession_id = d.accession_id
            WHERE d.status = 'Damaged'
            UNION ALL

            SELECT
                p.pcname AS `Item / PC Name`,
                'PC' AS Category,
                dept.department_name AS Facilities,
                p.location AS Location,
                p.accountable AS Accountable,
                COALESCE(dr.damage_type, 'Minor') AS Severity,
                dr.description AS `Damage Description`,
                p.acquisition_cost AS `Acquisition Cost`,
                p.date_acquired AS `Date Acquired`
            FROM pcinfofull p
            LEFT JOIN departments dept ON dept.department_id = p.department_id
            LEFT JOIN damage_reports dr ON dr.pcid = p.pcid
            WHERE p.status = 'Damaged'
        """

        filters = []
        params = []

        if name:
            filters.append("`Item / PC Name` LIKE %s")
            params.append(f"%{name}%")
        if category and category.lower() != "all":
            filters.append("Category = %s")
            params.append(category)
        if department and department.lower() != "all":
            filters.append("Facilities = %s")
            params.append(department)
        if severity and severity.lower() != "all":
            filters.append("Severity = %s")
            params.append(severity)
        if date_from:
            filters.append("`Date Acquired` >= %s")
            params.append(date_from)
        if date_to:
            filters.append("`Date Acquired` <= %s")
            params.append(date_to)

        if filters:
            query = f"SELECT * FROM ({query}) AS combined WHERE {' AND '.join(filters)}"

        query += " ORDER BY `Date Acquired` DESC"

        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(query, params)
            data = cursor.fetchall()

        df = pd.DataFrame(data)

        wb = Workbook()
        ws = wb.active
        ws.title = "Damage Reports"

        # Colour palette
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font = Font(name="Calibri", bold=True, color="1F4E79", size=16)
        subtitle_font = Font(name="Calibri", bold=False, color="4472C4", size=11)
        data_font = Font(name="Calibri", size=11, color="333333")
        data_font_bold = Font(name="Calibri", size=11, color="333333", bold=True)
        even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7"),
        )
        header_border = Border(
            left=Side(style="thin", color="1F4E79"),
            right=Side(style="thin", color="1F4E79"),
            top=Side(style="medium", color="1F4E79"),
            bottom=Side(style="medium", color="1F4E79"),
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1, value="Norzagaray College — IT Request Management System")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # Subtitle row
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
        subtitle_cell = ws.cell(row=2, column=1, value=f"Damage Report — Exported {datetime.now().strftime('%B %d, %Y %I:%M %p')}")
        subtitle_cell.font = subtitle_font
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # Blank spacer row
        ws.row_dimensions[3].height = 8

        # Header row (row 4)
        header_row = 4
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = header_border
        ws.row_dimensions[header_row].height = 28

        # Data rows
        currency_cols = {"Acquisition Cost"}
        date_cols = {"Date Acquired"}
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=header_row + 1):
            is_even = (row_idx - header_row) % 2 == 0
            row_fill = even_fill if is_even else odd_fill
            ws.row_dimensions[row_idx].height = 22

            for col_idx, col_name in enumerate(df.columns, start=1):
                val = row_data[col_name]
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border

                if col_name in currency_cols:
                    cell.number_format = '₱#,##0.00'
                    cell.alignment = right_align
                elif col_name in date_cols:
                    cell.number_format = 'YYYY-MM-DD'
                    cell.alignment = center_align
                elif col_name == "Severity":
                    cell.alignment = center_align
                    sev_val = str(val).strip().lower()
                    if sev_val == "critical":
                        cell.font = Font(name="Calibri", size=11, color="991B1B", bold=True)
                    elif sev_val == "major":
                        cell.font = Font(name="Calibri", size=11, color="9A3412", bold=True)
                    elif sev_val == "minor":
                        cell.font = Font(name="Calibri", size=11, color="854D0E", bold=True)
                elif col_name == "Item / PC Name":
                    cell.font = data_font_bold
                    cell.alignment = left_align
                else:
                    cell.alignment = left_align

        # Auto-width columns
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name)) + 2
            for row_idx in range(header_row + 1, header_row + 1 + len(df)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)

        # Freeze panes
        ws.freeze_panes = "A5"

        # Print settings
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = f"{header_row}:{header_row}"

        # Auto-filter
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(df.columns))}{header_row + len(df)}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            download_name="Damage_Reports.xlsx",
            as_attachment=True
        )
    except Exception as e:
        print(f"❌ Damage report export error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass