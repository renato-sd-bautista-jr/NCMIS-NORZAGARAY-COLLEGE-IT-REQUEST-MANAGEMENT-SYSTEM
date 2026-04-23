from flask import Blueprint, render_template, session, redirect, url_for, send_file, request
import datetime
import pymysql
from db import get_db_connection
from io import BytesIO

# Optional dependency: openpyxl is used to generate styled Excel files.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except Exception:
    Workbook = None

ticket_reports_bp = Blueprint('ticket_reports_bp', __name__, template_folder='templates')


@ticket_reports_bp.route('/tickets/history')
def ticket_history():
    user = session.get('user')
    if not user:
        return redirect(url_for('login_bp.login'))

    perms = user.get('permissions') or {}
    if isinstance(perms, str):
        try:
            import json
            perms = json.loads(perms)
        except Exception:
            perms = {}

    can_view_all = bool(user.get('is_admin')) or bool(perms.get('submit_ticket', {}).get('view'))

    conn = get_db_connection()
    tickets = []
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            # detect if ticket_responses exists so we can safely compute resolved_at
            try:
                cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_responses'")
                has_tr = bool(cursor.fetchone().get('cnt'))
            except Exception:
                has_tr = False

            # Also detect if tickets table has resolved_at/resolved_by_id columns (added by migrations)
            try:
                cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'resolved_at'")
                has_resolved_col = bool(cursor.fetchone().get('cnt'))
            except Exception:
                has_resolved_col = False
            try:
                cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'resolved_by_id'")
                has_resolved_by_col = bool(cursor.fetchone().get('cnt'))
            except Exception:
                has_resolved_by_col = False

            # Compute resolved_at expression: prefer explicit tickets.resolved_at when present,
            # otherwise fall back to ticket_responses subquery when that table exists.
            if has_resolved_col and has_tr:
                resolved_expr = "COALESCE(t.resolved_at, (SELECT MAX(tr.created_at) FROM ticket_responses tr WHERE tr.ticket_id = t.ticket_id AND tr.action = 'resolve')) AS resolved_at"
            elif has_resolved_col and not has_tr:
                resolved_expr = "t.resolved_at AS resolved_at"
            elif has_tr and not has_resolved_col:
                resolved_expr = "(SELECT MAX(tr.created_at) FROM ticket_responses tr WHERE tr.ticket_id = t.ticket_id AND tr.action = 'resolve') AS resolved_at"
            else:
                resolved_expr = "NULL AS resolved_at"

            # Build select, including resolved_by_id when available
            select_extra = ", t.resolved_by_id" if has_resolved_by_col else ""
            base_select = (
                "SELECT t.ticket_id, t.title, t.description, t.status, t.created_at, " + resolved_expr + ", t.user_id" + select_extra + ", "
                "COALESCE(CONCAT_WS(' ', u.first_name, u.last_name), u.faculty_name, u.username) AS submitter_name, d.department_name AS facility_name "
                "FROM tickets t "
                "LEFT JOIN departments d ON d.department_id = t.department_id "
                "LEFT JOIN users u ON u.user_id = t.user_id "
            )
            if can_view_all:
                cursor.execute(base_select + " ORDER BY t.created_at DESC")
            else:
                cursor.execute(base_select + " WHERE t.user_id=%s ORDER BY t.created_at DESC", (user.get('user_id'),))
            tickets = cursor.fetchall() or []

            # Attach resolver name (who performed 'resolve' action) for each ticket
            resolved_by_map = {}
            ticket_ids = [t['ticket_id'] for t in tickets]
            if ticket_ids and has_tr:
                try:
                    placeholders = ','.join(['%s'] * len(ticket_ids))
                    q = ("SELECT tr.ticket_id, tr.responder_id, tr.created_at, COALESCE(CONCAT_WS(' ', u.first_name, u.last_name), u.faculty_name, u.username) AS resolver_name "
                         "FROM ticket_responses tr LEFT JOIN users u ON u.user_id = tr.responder_id "
                         f"WHERE tr.ticket_id IN ({placeholders}) AND tr.action = 'resolve' ORDER BY tr.created_at DESC")
                    cursor.execute(q, ticket_ids)
                    resolve_rows = cursor.fetchall() or []
                    for r in resolve_rows:
                        tid = r['ticket_id']
                        if tid not in resolved_by_map:
                            resolved_by_map[tid] = r
                except Exception:
                    resolved_by_map = {}
                    # If tickets table stores resolved_by_id directly (migration), fetch those user names too
                    if has_resolved_by_col:
                        try:
                            # collect pairs of ticket_id -> resolved_by_id from fetched tickets
                            ticket_resolver_map = {t['ticket_id']: t.get('resolved_by_id') for t in tickets if t.get('resolved_by_id')}
                            resolver_ids = sorted({v for v in ticket_resolver_map.values() if v})
                            if resolver_ids:
                                placeholders_u = ','.join(['%s'] * len(resolver_ids))
                                cursor.execute(f"SELECT user_id, CONCAT_WS(' ', first_name, last_name) AS name, faculty_name, username FROM users WHERE user_id IN ({placeholders_u})", resolver_ids)
                                users = {u['user_id']: (u.get('name') or u.get('faculty_name') or u.get('username')) for u in (cursor.fetchall() or [])}
                                for tid, rid in ticket_resolver_map.items():
                                    if tid not in resolved_by_map and rid:
                                        resolved_by_map[tid] = {'responder_id': rid, 'resolver_name': users.get(rid) or None, 'created_at': next((tt.get('resolved_at') for tt in tickets if tt.get('ticket_id')==tid), None)}
                        except Exception:
                            pass
    finally:
        conn.close()

    # Attach helper fields onto tickets for the template (resolver fields)
    try:
        for t in tickets:
            tid = t.get('ticket_id')
            rb = resolved_by_map.get(tid) if 'resolved_by_map' in locals() else None
            t['resolved_by'] = rb.get('resolver_name') if rb else None
            t['resolved_by_id'] = rb.get('responder_id') if rb else None
            # prefer explicit resolved_at from resolved_by_map if available
            if rb and rb.get('created_at'):
                t['resolved_at'] = rb.get('created_at')
    except Exception:
        pass

    # Compute summary counts for the history view
    resolved = unresolved = declined = 0
    try:
        for t in (tickets or []):
            st = (t.get('status') or '')
            st_key = st.strip().lower() if isinstance(st, str) else str(st).strip().lower()
            if st_key in ('resolved', 'closed'):
                resolved += 1
            elif st_key in ('declined', 'rejected', 'invalid'):
                declined += 1
            elif st_key in ('unresolved',):
                unresolved += 1
    except Exception:
        resolved = unresolved = declined = 0

    # If ticket_responses exists, use it to supplement resolved count
    try:
        conn3 = get_db_connection()
        try:
            with conn3.cursor(pymysql.cursors.DictCursor) as cursor:
                try:
                    cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_responses'")
                    has_tr = bool(cursor.fetchone().get('cnt'))
                except Exception:
                    has_tr = False

                if has_tr:
                    try:
                        cursor.execute("SELECT COUNT(DISTINCT ticket_id) AS cnt FROM ticket_responses WHERE action = 'resolve'")
                        row = cursor.fetchone() or {}
                        resolved_from_responses = int(row.get('cnt') or 0)
                        if resolved_from_responses > resolved:
                            resolved = resolved_from_responses
                    except Exception:
                        pass
        finally:
            conn3.close()
    except Exception:
        pass

    try:
        status_keys = sorted({((t.get('status') or '') if t.get('status') is not None else '').strip() for t in (tickets or [])})
    except Exception:
        status_keys = []

    # Debug output to help diagnose counting issues (will appear in server console)
    try:
        print(f"TICKET_HISTORY DEBUG: tickets_count={len(tickets)} resolved={resolved} unresolved={unresolved} declined={declined}")
        print(f"TICKET_HISTORY DEBUG: status_keys={status_keys}")
        for t in (tickets or [])[:10]:
            print(f"TICKET_HISTORY DEBUG: ticket_id={t.get('ticket_id')} status={repr(t.get('status'))} resolved_at={t.get('resolved_at')} resolved_by={t.get('resolved_by')}")
    except Exception:
        pass

    return render_template('ticket_history.html', tickets=tickets, resolved=resolved, unresolved=unresolved, declined=declined, status_keys=status_keys)


@ticket_reports_bp.route('/tickets/history/export')
def export_ticket_history_excel():
    """Export ticket history as a styled Excel workbook."""
    user = session.get('user')
    if not user:
        return redirect(url_for('login_bp.login'))

    if Workbook is None:
        return "openpyxl is not installed. Please install openpyxl to enable Excel export.", 500

    perms = user.get('permissions') or {}
    if isinstance(perms, str):
        try:
            import json
            perms = json.loads(perms)
        except Exception:
            perms = {}

    can_view_all = bool(user.get('is_admin')) or bool(perms.get('submit_ticket', {}).get('view'))

    # parse filter query params so the exported workbook can match the UI table
    q = (request.args.get('q') or '').strip()
    facility = (request.args.get('facility') or '').strip()
    device = (request.args.get('device') or '').strip()
    status = (request.args.get('status') or '').strip()
    timeframe = (request.args.get('timeframe') or '').strip()

    conn = get_db_connection()
    tickets = []
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            # detect if ticket_responses exists so we can compute resolved_at safely
            try:
                cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_responses'")
                has_tr = bool(cursor.fetchone().get('cnt'))
            except Exception:
                has_tr = False

            # Determine resolved_at/resolved_by availability
            try:
                cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'resolved_at'")
                has_resolved_col = bool(cursor.fetchone().get('cnt'))
            except Exception:
                has_resolved_col = False
            try:
                cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'resolved_by_id'")
                has_resolved_by_col = bool(cursor.fetchone().get('cnt'))
            except Exception:
                has_resolved_by_col = False

            if has_resolved_col and has_tr:
                resolved_expr = "COALESCE(t.resolved_at, (SELECT MAX(tr.created_at) FROM ticket_responses tr WHERE tr.ticket_id = t.ticket_id AND tr.action = 'resolve')) AS resolved_at"
            elif has_resolved_col and not has_tr:
                resolved_expr = "t.resolved_at AS resolved_at"
            elif has_tr and not has_resolved_col:
                resolved_expr = "(SELECT MAX(tr.created_at) FROM ticket_responses tr WHERE tr.ticket_id = t.ticket_id AND tr.action = 'resolve') AS resolved_at"
            else:
                resolved_expr = "NULL AS resolved_at"

            select_extra = ", t.resolved_by_id" if has_resolved_by_col else ""
            # build base select + dynamic where clauses
            base_select = (
                "SELECT t.ticket_id, t.user_id, t.department_id, t.title, t.description, t.status, t.created_at, " + resolved_expr + ","
                " COALESCE(CONCAT_WS(' ', u.first_name, u.last_name), u.faculty_name, u.username) AS submitter_name, d.department_name AS facility_name, t.device_id" + select_extra + " "
                "FROM tickets t "
                "LEFT JOIN users u ON u.user_id = t.user_id "
                "LEFT JOIN departments d ON d.department_id = t.department_id "
            )

            where_clauses = []
            params = []

            if not can_view_all:
                where_clauses.append('t.user_id=%s')
                params.append(user.get('user_id'))

            if status:
                where_clauses.append('LOWER(t.status)=%s')
                params.append(status.lower())

            if q:
                like = f"%{q}%"
                where_clauses.append("(t.title LIKE %s OR t.description LIKE %s OR COALESCE(CONCAT_WS(' ', u.first_name, u.last_name), u.faculty_name, u.username) LIKE %s OR u.email LIKE %s)")
                params.extend([like, like, like, like])

            # timeframe filter: support 'Last 30 days' and 'Last 90 days'
            if timeframe:
                try:
                    if '30' in timeframe:
                        cutoff = datetime.datetime.utcnow() - datetime.timedelta(days=30)
                    elif '90' in timeframe:
                        cutoff = datetime.datetime.utcnow() - datetime.timedelta(days=90)
                    else:
                        cutoff = None
                    if cutoff:
                        where_clauses.append('t.created_at >= %s')
                        params.append(cutoff.strftime('%Y-%m-%d %H:%M:%S'))
                except Exception:
                    pass

            # assemble initial query
            if where_clauses:
                qsql = base_select + ' WHERE ' + ' AND '.join(where_clauses) + ' ORDER BY t.created_at DESC'
                cursor.execute(qsql, params)
            else:
                cursor.execute(base_select + ' ORDER BY t.created_at DESC')

            tickets = cursor.fetchall() or []

            ticket_ids = [t['ticket_id'] for t in tickets]

            # Attach latest response for each ticket (if ticket_responses exists)
            latest_responses = {}
            if ticket_ids:
                placeholders = ','.join(['%s'] * len(ticket_ids))
                try:
                    cursor.execute(f"SELECT ticket_id, feedback, responder_id, created_at FROM ticket_responses WHERE ticket_id IN ({placeholders}) ORDER BY created_at DESC", ticket_ids)
                    resp_rows = cursor.fetchall() or []
                    for r in resp_rows:
                        tid = r['ticket_id']
                        if tid not in latest_responses:
                            latest_responses[tid] = r
                except Exception:
                    latest_responses = {}

            # Attach device names (support ticket_devices or tickets.device_id)
            devices_map = {}
            if ticket_ids:
                try:
                    cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_devices'")
                    has_td = bool(cursor.fetchone().get('cnt'))
                except Exception:
                    has_td = False

                if has_td:
                    try:
                        placeholders = ','.join(['%s'] * len(ticket_ids))
                        q2 = ("SELECT td.ticket_id, COALESCE(df.item_name, d.item_name) AS device_name "
                             "FROM ticket_devices td "
                             "LEFT JOIN devices_full df ON df.accession_id = td.device_id "
                             "LEFT JOIN devices d ON d.device_id = td.device_id "
                             f"WHERE td.ticket_id IN ({placeholders})")
                        cursor.execute(q2, ticket_ids)
                        assoc = cursor.fetchall() or []
                        for a in assoc:
                            devices_map.setdefault(a['ticket_id'], []).append(a.get('device_name'))
                    except Exception:
                        devices_map = {}
                else:
                    # fallback to tickets.device_id column
                    try:
                        cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'device_id'")
                        has_devcol = bool(cursor.fetchone().get('cnt'))
                    except Exception:
                        has_devcol = False
                    if has_devcol:
                        dids = [t.get('device_id') for t in tickets if t.get('device_id')]
                        if dids:
                            placeholders = ','.join(['%s'] * len(dids))
                            q3 = ("SELECT COALESCE(df.accession_id, d.device_id) AS id, COALESCE(df.item_name, d.item_name) AS name "
                                 "FROM devices_full df LEFT JOIN devices d ON d.device_id = df.accession_id "
                                 f"WHERE COALESCE(df.accession_id, d.device_id) IN ({placeholders})")
                            cursor.execute(q3, dids)
                            devrows = cursor.fetchall() or []
                            devmap = {dr['id']: dr['name'] for dr in devrows}
                            for t in tickets:
                                did = t.get('device_id')
                                if did:
                                    devices_map.setdefault(t['ticket_id'], []).append(devmap.get(did))
                    else:
                        for t in tickets:
                            devices_map.setdefault(t['ticket_id'], [])

            # if a device filter was provided, further filter the tickets list in Python
            if device:
                try:
                    # device param may be id; filter tickets that have that device
                    filtered = []
                    for t in tickets:
                        tid = t.get('ticket_id')
                        devs = devices_map.get(tid, [])
                        # check id in original device_id or name match
                        if str(t.get('device_id') or '') == str(device) or any(str(device) in str(d) for d in devs if d):
                            filtered.append(t)
                    tickets = filtered
                except Exception:
                    pass

            # if a facility filter was provided, further filter tickets by department id or name
            if facility:
                try:
                    filtered = []
                    for t in tickets:
                        if str(t.get('department_id') or '') == str(facility) or str(t.get('facility_name') or '').lower().find(facility.lower()) != -1:
                            filtered.append(t)
                    tickets = filtered
                except Exception:
                    pass

    finally:
        conn.close()

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ticket History'

    headers = ['Ticket ID', 'User ID', 'Submitter', 'Department', 'Title', 'Description', 'Devices', 'Status', 'Created At', 'Resolved At', 'Latest Response', 'Response At', 'Responder ID']

    header_font = Font(bold=True, color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF4F46E5', end_color='FF4F46E5', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    wrap = Alignment(wrap_text=True)
    thin = Side(border_style='thin', color='FFBBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    # populate rows from tickets
    row_idx = 2
    for t in tickets:
        tid = t.get('ticket_id')
        devices = '; '.join([d for d in (devices_map.get(tid) or []) if d])
        lr = latest_responses.get(tid) or {}
        ws.cell(row=row_idx, column=1, value=tid)
        ws.cell(row=row_idx, column=2, value=t.get('user_id'))
        ws.cell(row=row_idx, column=3, value=t.get('submitter_name'))
        ws.cell(row=row_idx, column=4, value=t.get('facility_name'))
        ws.cell(row=row_idx, column=5, value=t.get('title'))
        desc_cell = ws.cell(row=row_idx, column=6, value=t.get('description'))
        desc_cell.alignment = wrap
        ws.cell(row=row_idx, column=7, value=devices)
        ws.cell(row=row_idx, column=8, value=t.get('status'))
        ws.cell(row=row_idx, column=9, value=str(t.get('created_at')))
        ws.cell(row=row_idx, column=10, value=str(t.get('resolved_at')) if t.get('resolved_at') else '')
        ws.cell(row=row_idx, column=11, value=lr.get('feedback'))
        ws.cell(row=row_idx, column=12, value=str(lr.get('created_at')) if lr.get('created_at') else '')
        ws.cell(row=row_idx, column=13, value=lr.get('responder_id'))
        row_idx += 1

    # adjust column widths for readability
    widths = [12, 12, 24, 20, 30, 60, 30, 16, 20, 20, 40, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # create output for history export
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='ticket_history.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@ticket_reports_bp.route('/tickets/report')
def ticket_report():
    user = session.get('user')
    if not user:
        return redirect(url_for('login_bp.login'))

    # Fetch status aggregates
    conn = get_db_connection()
    stats = []
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            try:
                cursor.execute("SELECT status, COUNT(*) AS cnt FROM tickets GROUP BY status")
                stats = cursor.fetchall() or []
            except Exception:
                stats = []
    finally:
        conn.close()

    # Compute aggregates server-side to avoid template scoping issues
    resolved = 0
    unresolved = 0
    declined = 0
    try:
        for r in stats:
            st = (r.get('status') or '').strip().lower()
            try:
                cnt = int(r.get('cnt') or 0)
            except Exception:
                try:
                    cnt = int(float(r.get('cnt')))
                except Exception:
                    cnt = 0

            if st in ('resolved', 'closed'):
                resolved += cnt
            elif st in ('declined', 'rejected', 'invalid'):
                declined += cnt
            elif st in ('unresolved',):
                unresolved += cnt
    except Exception:
        resolved = unresolved = declined = 0

    # If ticket_responses table exists, prefer counting tickets with a 'resolve' response
    conn3 = get_db_connection()
    try:
        with conn3.cursor(pymysql.cursors.DictCursor) as cursor:
            try:
                cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_responses'")
                has_tr = bool(cursor.fetchone().get('cnt'))
            except Exception:
                has_tr = False

            if has_tr:
                try:
                    cursor.execute("SELECT COUNT(DISTINCT ticket_id) AS cnt FROM ticket_responses WHERE action = 'resolve'")
                    row = cursor.fetchone() or {}
                    resolved_from_responses = int(row.get('cnt') or 0)
                    # Use the larger of the two counts to avoid undercounting when status wasn't updated
                    if resolved_from_responses > resolved:
                        resolved = resolved_from_responses
                except Exception:
                    pass
    finally:
        conn3.close()

    # Also fetch full ticket rows and related info to show details in the report
    tickets = []
    try:
        perms = user.get('permissions') or {}
        if isinstance(perms, str):
            try:
                import json
                perms = json.loads(perms)
            except Exception:
                perms = {}

        can_view_all = bool(user.get('is_admin')) or bool(perms.get('submit_ticket', {}).get('view'))

        conn2 = get_db_connection()
        try:
            with conn2.cursor(pymysql.cursors.DictCursor) as cursor:
                try:
                    cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_responses'")
                    has_tr = bool(cursor.fetchone().get('cnt'))
                except Exception:
                    has_tr = False

                try:
                    cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'resolved_at'")
                    has_resolved_col = bool(cursor.fetchone().get('cnt'))
                except Exception:
                    has_resolved_col = False
                try:
                    cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'resolved_by_id'")
                    has_resolved_by_col = bool(cursor.fetchone().get('cnt'))
                except Exception:
                    has_resolved_by_col = False

                if has_resolved_col and has_tr:
                    resolved_expr = "COALESCE(t.resolved_at, (SELECT MAX(tr.created_at) FROM ticket_responses tr WHERE tr.ticket_id = t.ticket_id AND tr.action = 'resolve')) AS resolved_at"
                elif has_resolved_col and not has_tr:
                    resolved_expr = "t.resolved_at AS resolved_at"
                elif has_tr and not has_resolved_col:
                    resolved_expr = "(SELECT MAX(tr.created_at) FROM ticket_responses tr WHERE tr.ticket_id = t.ticket_id AND tr.action = 'resolve') AS resolved_at"
                else:
                    resolved_expr = "NULL AS resolved_at"

                select_extra = ", t.resolved_by_id" if has_resolved_by_col else ""

                base_select = (
                    "SELECT t.ticket_id, t.user_id, t.department_id, t.title, t.description, t.status, t.created_at, " + resolved_expr + ","
                    " COALESCE(CONCAT_WS(' ', u.first_name, u.last_name), u.faculty_name, u.username) AS submitter_name, d.department_name AS facility_name, t.device_id" + select_extra + " "
                    "FROM tickets t "
                    "LEFT JOIN users u ON u.user_id = t.user_id "
                    "LEFT JOIN departments d ON d.department_id = t.department_id "
                )

                if can_view_all:
                    cursor.execute(base_select + " ORDER BY t.created_at DESC")
                else:
                    cursor.execute(base_select + " WHERE t.user_id=%s ORDER BY t.created_at DESC", (user.get('user_id'),))

                tickets = cursor.fetchall() or []

                ticket_ids = [t['ticket_id'] for t in tickets]

                # Attach latest response for each ticket (if ticket_responses exists)
                latest_responses = {}
                if ticket_ids:
                    placeholders = ','.join(['%s'] * len(ticket_ids))
                    try:
                        cursor.execute(f"SELECT ticket_id, feedback, responder_id, created_at FROM ticket_responses WHERE ticket_id IN ({placeholders}) ORDER BY created_at DESC", ticket_ids)
                        resp_rows = cursor.fetchall() or []
                        for r in resp_rows:
                            tid = r['ticket_id']
                            if tid not in latest_responses:
                                latest_responses[tid] = r
                    except Exception:
                        latest_responses = {}

                # Attach resolver name (who performed 'resolve' action) for each ticket
                resolved_by_map = {}
                if ticket_ids and has_tr:
                    try:
                        placeholders = ','.join(['%s'] * len(ticket_ids))
                        q = ("SELECT tr.ticket_id, tr.responder_id, tr.created_at, COALESCE(CONCAT_WS(' ', u.first_name, u.last_name), u.faculty_name, u.username) AS resolver_name "
                             "FROM ticket_responses tr LEFT JOIN users u ON u.user_id = tr.responder_id "
                             f"WHERE tr.ticket_id IN ({placeholders}) AND tr.action = 'resolve' ORDER BY tr.created_at DESC")
                        cursor.execute(q, ticket_ids)
                        resolve_rows = cursor.fetchall() or []
                        for r in resolve_rows:
                            tid = r['ticket_id']
                            if tid not in resolved_by_map:
                                resolved_by_map[tid] = {
                                    'responder_id': r.get('responder_id'),
                                    'resolved_at': r.get('created_at'),
                                    'resolver_name': r.get('resolver_name')
                                }
                    except Exception:
                        resolved_by_map = {}
                # If tickets table stores resolved_by_id directly (migration), fetch those user names too
                if 'has_resolved_by_col' in locals() and has_resolved_by_col:
                    try:
                        ticket_resolver_map = {t['ticket_id']: t.get('resolved_by_id') for t in tickets if t.get('resolved_by_id')}
                        resolver_ids = sorted({v for v in ticket_resolver_map.values() if v})
                        if resolver_ids:
                            placeholders_u = ','.join(['%s'] * len(resolver_ids))
                            cursor.execute(f"SELECT user_id, CONCAT_WS(' ', first_name, last_name) AS name, faculty_name, username FROM users WHERE user_id IN ({placeholders_u})", resolver_ids)
                            users = {u['user_id']: (u.get('name') or u.get('faculty_name') or u.get('username')) for u in (cursor.fetchall() or [])}
                            for tid, rid in ticket_resolver_map.items():
                                if tid not in resolved_by_map and rid:
                                    resolved_by_map[tid] = {'responder_id': rid, 'resolver_name': users.get(rid) or None, 'resolved_at': next((tt.get('resolved_at') for tt in tickets if tt.get('ticket_id')==tid), None)}
                    except Exception:
                        pass

                # Attach device names (support ticket_devices or tickets.device_id)
                devices_map = {}
                if ticket_ids:
                    try:
                        cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_devices'")
                        has_td = bool(cursor.fetchone().get('cnt'))
                    except Exception:
                        has_td = False

                    if has_td:
                        try:
                            placeholders = ','.join(['%s'] * len(ticket_ids))
                            q = ("SELECT td.ticket_id, COALESCE(df.item_name, d.item_name) AS device_name "
                                 "FROM ticket_devices td "
                                 "LEFT JOIN devices_full df ON df.accession_id = td.device_id "
                                 "LEFT JOIN devices d ON d.device_id = td.device_id "
                                 f"WHERE td.ticket_id IN ({placeholders})")
                            cursor.execute(q, ticket_ids)
                            assoc = cursor.fetchall() or []
                            for a in assoc:
                                devices_map.setdefault(a['ticket_id'], []).append(a.get('device_name'))
                        except Exception:
                            devices_map = {}
                    else:
                        try:
                            cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'device_id'")
                            has_devcol = bool(cursor.fetchone().get('cnt'))
                        except Exception:
                            has_devcol = False
                        if has_devcol:
                            dids = [t.get('device_id') for t in tickets if t.get('device_id')]
                            if dids:
                                placeholders = ','.join(['%s'] * len(dids))
                                q = ("SELECT COALESCE(df.accession_id, d.device_id) AS id, COALESCE(df.item_name, d.item_name) AS name "
                                     "FROM devices_full df LEFT JOIN devices d ON d.device_id = df.accession_id "
                                     f"WHERE COALESCE(df.accession_id, d.device_id) IN ({placeholders})")
                                cursor.execute(q, dids)
                                devrows = cursor.fetchall() or []
                                devmap = {dr['id']: dr['name'] for dr in devrows}
                                for t in tickets:
                                    did = t.get('device_id')
                                    if did:
                                        devices_map.setdefault(t['ticket_id'], []).append(devmap.get(did))
                        else:
                            for t in tickets:
                                devices_map.setdefault(t['ticket_id'], [])

                # Attach helper fields onto tickets for the template
                for t in tickets:
                    tid = t.get('ticket_id')
                    t['devices'] = devices_map.get(tid, [])
                    lr = latest_responses.get(tid) or {}
                    t['latest_response'] = lr.get('feedback')
                    t['response_at'] = lr.get('created_at')
                    t['responder_id'] = lr.get('responder_id')
                    # resolved_by_map may contain the admin who marked the ticket resolved
                    rb = resolved_by_map.get(tid) if 'resolved_by_map' in locals() else None
                    t['resolved_by'] = rb.get('resolver_name') if rb else None
                    t['resolved_by_id'] = rb.get('responder_id') if rb else None
                    # prefer explicit resolved_at from resolved_by_map if available
                    if rb and rb.get('resolved_at'):
                        t['resolved_at'] = rb.get('resolved_at')
        finally:
            conn2.close()
    except Exception:
        tickets = []

    return render_template('ticket_report.html', stats=stats, resolved=resolved, unresolved=unresolved, declined=declined, tickets=tickets)
 


@ticket_reports_bp.route('/tickets/report/export')
def export_ticket_report_excel():
    """Export ticket report counts as a styled Excel workbook."""
    user = session.get('user')
    if not user:
        return redirect(url_for('login_bp.login'))

    if Workbook is None:
        return "openpyxl is not installed. Please install openpyxl to enable Excel export.", 500

    conn = get_db_connection()
    stats = []
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            try:
                cursor.execute("SELECT status, COUNT(*) AS cnt FROM tickets GROUP BY status")
                stats = cursor.fetchall() or []
            except Exception:
                stats = []
    finally:
        conn.close()

    # Also fetch full ticket rows to export the "All Tickets" table
    tickets = []
    try:
        perms = user.get('permissions') or {}
        if isinstance(perms, str):
            try:
                import json
                perms = json.loads(perms)
            except Exception:
                perms = {}

        can_view_all = bool(user.get('is_admin')) or bool(perms.get('submit_ticket', {}).get('view'))

        conn2 = get_db_connection()
        try:
            with conn2.cursor(pymysql.cursors.DictCursor) as cursor:
                try:
                    cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_responses'")
                    has_tr = bool(cursor.fetchone().get('cnt'))
                except Exception:
                    has_tr = False

                try:
                    cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'resolved_at'")
                    has_resolved_col = bool(cursor.fetchone().get('cnt'))
                except Exception:
                    has_resolved_col = False
                try:
                    cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'resolved_by_id'")
                    has_resolved_by_col = bool(cursor.fetchone().get('cnt'))
                except Exception:
                    has_resolved_by_col = False

                if has_resolved_col and has_tr:
                    resolved_expr = "COALESCE(t.resolved_at, (SELECT MAX(tr.created_at) FROM ticket_responses tr WHERE tr.ticket_id = t.ticket_id AND tr.action = 'resolve')) AS resolved_at"
                elif has_resolved_col and not has_tr:
                    resolved_expr = "t.resolved_at AS resolved_at"
                elif has_tr and not has_resolved_col:
                    resolved_expr = "(SELECT MAX(tr.created_at) FROM ticket_responses tr WHERE tr.ticket_id = t.ticket_id AND tr.action = 'resolve') AS resolved_at"
                else:
                    resolved_expr = "NULL AS resolved_at"

                select_extra = ", t.resolved_by_id" if has_resolved_by_col else ""

                base_select = (
                    "SELECT t.ticket_id, t.user_id, t.department_id, t.title, t.description, t.status, t.created_at, " + resolved_expr + ","
                    " COALESCE(CONCAT_WS(' ', u.first_name, u.last_name), u.faculty_name, u.username) AS submitter_name, d.department_name AS facility_name, t.device_id" + select_extra + " "
                    "FROM tickets t "
                    "LEFT JOIN users u ON u.user_id = t.user_id "
                    "LEFT JOIN departments d ON d.department_id = t.department_id "
                )

                if can_view_all:
                    cursor.execute(base_select + " ORDER BY t.created_at DESC")
                else:
                    cursor.execute(base_select + " WHERE t.user_id=%s ORDER BY t.created_at DESC", (user.get('user_id'),))

                tickets = cursor.fetchall() or []

                ticket_ids = [t['ticket_id'] for t in tickets]

                # Attach latest response for each ticket (if ticket_responses exists)
                latest_responses = {}
                if ticket_ids:
                    placeholders = ','.join(['%s'] * len(ticket_ids))
                    try:
                        cursor.execute(f"SELECT ticket_id, feedback, responder_id, created_at FROM ticket_responses WHERE ticket_id IN ({placeholders}) ORDER BY created_at DESC", ticket_ids)
                        resp_rows = cursor.fetchall() or []
                        for r in resp_rows:
                            tid = r['ticket_id']
                            if tid not in latest_responses:
                                latest_responses[tid] = r
                    except Exception:
                        latest_responses = {}

                # Attach resolver name (who performed 'resolve' action) for each ticket
                resolved_by_map = {}
                if ticket_ids and has_tr:
                    try:
                        placeholders = ','.join(['%s'] * len(ticket_ids))
                        q = ("SELECT tr.ticket_id, tr.responder_id, tr.created_at, COALESCE(CONCAT_WS(' ', u.first_name, u.last_name), u.faculty_name, u.username) AS resolver_name "
                             "FROM ticket_responses tr LEFT JOIN users u ON u.user_id = tr.responder_id "
                             f"WHERE tr.ticket_id IN ({placeholders}) AND tr.action = 'resolve' ORDER BY tr.created_at DESC")
                        cursor.execute(q, ticket_ids)
                        resolve_rows = cursor.fetchall() or []
                        for r in resolve_rows:
                            tid = r['ticket_id']
                            if tid not in resolved_by_map:
                                resolved_by_map[tid] = {
                                    'responder_id': r.get('responder_id'),
                                    'resolved_at': r.get('created_at'),
                                    'resolver_name': r.get('resolver_name')
                                }
                    except Exception:
                        resolved_by_map = {}
                # If tickets table stores resolved_by_id directly (migration), fetch those user names too
                if 'has_resolved_by_col' in locals() and has_resolved_by_col:
                    try:
                        ticket_resolver_map = {t['ticket_id']: t.get('resolved_by_id') for t in tickets if t.get('resolved_by_id')}
                        resolver_ids = sorted({v for v in ticket_resolver_map.values() if v})
                        if resolver_ids:
                            placeholders_u = ','.join(['%s'] * len(resolver_ids))
                            cursor.execute(f"SELECT user_id, CONCAT_WS(' ', first_name, last_name) AS name, faculty_name, username FROM users WHERE user_id IN ({placeholders_u})", resolver_ids)
                            users = {u['user_id']: (u.get('name') or u.get('faculty_name') or u.get('username')) for u in (cursor.fetchall() or [])}
                            for tid, rid in ticket_resolver_map.items():
                                if tid not in resolved_by_map and rid:
                                    resolved_by_map[tid] = {'responder_id': rid, 'resolver_name': users.get(rid) or None, 'resolved_at': next((tt.get('resolved_at') for tt in tickets if tt.get('ticket_id')==tid), None)}
                    except Exception:
                        pass

                # Attach device names (support ticket_devices or tickets.device_id)
                devices_map = {}
                if ticket_ids:
                    try:
                        cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'ticket_devices'")
                        has_td = bool(cursor.fetchone().get('cnt'))
                    except Exception:
                        has_td = False

                    if has_td:
                        try:
                            placeholders = ','.join(['%s'] * len(ticket_ids))
                            q = ("SELECT td.ticket_id, COALESCE(df.item_name, d.item_name) AS device_name "
                                 "FROM ticket_devices td "
                                 "LEFT JOIN devices_full df ON df.accession_id = td.device_id "
                                 "LEFT JOIN devices d ON d.device_id = td.device_id "
                                 f"WHERE td.ticket_id IN ({placeholders})")
                            cursor.execute(q, ticket_ids)
                            assoc = cursor.fetchall() or []
                            for a in assoc:
                                devices_map.setdefault(a['ticket_id'], []).append(a.get('device_name'))
                        except Exception:
                            devices_map = {}
                    else:
                        try:
                            cursor.execute("SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'tickets' AND COLUMN_NAME = 'device_id'")
                            has_devcol = bool(cursor.fetchone().get('cnt'))
                        except Exception:
                            has_devcol = False
                        if has_devcol:
                            dids = [t.get('device_id') for t in tickets if t.get('device_id')]
                            if dids:
                                placeholders = ','.join(['%s'] * len(dids))
                                q = ("SELECT COALESCE(df.accession_id, d.device_id) AS id, COALESCE(df.item_name, d.item_name) AS name "
                                     "FROM devices_full df LEFT JOIN devices d ON d.device_id = df.accession_id "
                                     f"WHERE COALESCE(df.accession_id, d.device_id) IN ({placeholders})")
                                cursor.execute(q, dids)
                                devrows = cursor.fetchall() or []
                                devmap = {dr['id']: dr['name'] for dr in devrows}
                                for t in tickets:
                                    did = t.get('device_id')
                                    if did:
                                        devices_map.setdefault(t['ticket_id'], []).append(devmap.get(did))
                        else:
                            for t in tickets:
                                devices_map.setdefault(t['ticket_id'], [])

                # Attach helper fields onto tickets for the export
                for t in tickets:
                    tid = t.get('ticket_id')
                    t['devices'] = devices_map.get(tid, [])
                    lr = latest_responses.get(tid) or {}
                    t['latest_response'] = lr.get('feedback')
                    t['response_at'] = lr.get('created_at')
                    t['responder_id'] = lr.get('responder_id')
                    rb = resolved_by_map.get(tid) if 'resolved_by_map' in locals() else None
                    t['resolved_by'] = rb.get('resolver_name') if rb else None
                    t['resolved_by_id'] = rb.get('responder_id') if rb else None
                    if rb and rb.get('resolved_at'):
                        t['resolved_at'] = rb.get('resolved_at')
        finally:
            conn2.close()
    except Exception:
        tickets = []

    # Build workbook with detailed tickets as the first sheet so it opens by default
    wb = Workbook()

    # Prepare shared styles
    header_font = Font(bold=True, color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF2563EB', end_color='FF2563EB', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    # Use the initially created active sheet for the detailed tickets
    ws2 = wb.active
    ws2.title = 'All Tickets'
    headers = ['Ticket ID', 'Title', 'Submitter', 'Department', 'Devices', 'Status', 'Created At', 'Resolved At', 'Resolved By']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row_idx = 2
    for t in tickets:
        devices = ', '.join([d for d in (t.get('devices') or []) if d])
        ws2.cell(row=row_idx, column=1, value=t.get('ticket_id'))
        ws2.cell(row=row_idx, column=2, value=t.get('title'))
        ws2.cell(row=row_idx, column=3, value=t.get('submitter_name'))
        ws2.cell(row=row_idx, column=4, value=t.get('facility_name'))
        ws2.cell(row=row_idx, column=5, value=devices)
        ws2.cell(row=row_idx, column=6, value=t.get('status'))
        ws2.cell(row=row_idx, column=7, value=str(t.get('created_at')))
        ws2.cell(row=row_idx, column=8, value=str(t.get('resolved_at')) if t.get('resolved_at') else '')
        ws2.cell(row=row_idx, column=9, value=t.get('resolved_by') or '')
        row_idx += 1

    # adjust column widths for detailed sheet
    widths = [12, 40, 24, 20, 30, 16, 20, 20, 24]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    # Add a second sheet for the summary counts
    ws = wb.create_sheet(title='Ticket Report')
    ws.cell(row=1, column=1, value='Status').font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2, value='Count').font = header_font
    ws.cell(row=1, column=2).fill = header_fill

    r = 2
    total = 0
    for s in stats:
        ws.cell(row=r, column=1, value=s.get('status'))
        ws.cell(row=r, column=2, value=s.get('cnt'))
        total += (s.get('cnt') or 0)
        r += 1

    ws.cell(row=r+1, column=1, value='Total').font = Font(bold=True)
    ws.cell(row=r+1, column=2, value=total).font = Font(bold=True)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='ticket_report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
